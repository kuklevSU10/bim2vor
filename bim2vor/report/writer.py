# -*- coding: utf-8 -*-
"""
Writer: заполняет шаблон ВОР Excel результатами специалистов и создаёт audit-отчёт.
Сохраняет форматирование оригинала.
"""
from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# Цвета для подсветки степени уверенности
FILL_OK = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_WARN = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_ALARM = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_NOT_IN_MODEL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_DELEGATED = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")


def confidence_to_fill(confidence: float, fill_status: str) -> PatternFill | None:
    if fill_status in ("not_in_model", "missing_in_ar_model", "not_in_bim_scope", "no_class_b_info_in_model"):
        return FILL_NOT_IN_MODEL
    if fill_status.startswith("delegated_"):
        return FILL_DELEGATED
    if confidence >= 0.65:
        return FILL_OK
    if confidence >= 0.4:
        return FILL_WARN
    return FILL_ALARM


def load_specialist_outputs(specialist_outputs_dir: Path) -> dict[str, dict]:
    """Загружает все JSON-результаты специалистов."""
    out: dict[str, dict] = {}
    for path in sorted(specialist_outputs_dir.glob("*.json")):
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            out[data.get("specialist", path.stem)] = data
        except Exception as e:
            print(f"  ! Не смог прочитать {path.name}: {e}")
    return out


def consolidate_allocations(
    specialist_outputs: dict[str, dict],
) -> dict[str, list[dict]]:
    """
    Объединяет аллокации от всех специалистов.
    position_id → [список аллокаций от разных специалистов]
    """
    by_position: dict[str, list[dict]] = defaultdict(list)
    for spec_key, data in specialist_outputs.items():
        for alloc in data.get("phase3_allocations", []):
            pos = alloc.get("position_id", "").strip().rstrip(".")
            if pos:
                alloc["_specialist"] = spec_key
                by_position[pos].append(alloc)
    return by_position


def _detect_write_columns(ws) -> dict[str, int]:
    """Автодетект колонок для записи результатов."""
    cols = {}
    for c in range(1, min(ws.max_column or 30, 30) + 1):
        v = ws.cell(1, c).value
        if not v:
            continue
        h = str(v).strip().lower()
        if "количество заказчика" in h or "количество подрядн" in h:
            cols["qty_planned"] = c
        elif "количество гп" in h:
            cols["qty_gp"] = c
        elif "примечание гп" in h:
            cols["gp_note"] = c
        elif "номер позиции" in h or "шифр позиции" in h:
            cols["code"] = c
    return cols


def fill_boq_template(
    template_path: Path,
    output_path: Path,
    by_position: dict[str, list[dict]],
    overall_run_summary: dict | None = None,
) -> dict:
    """
    Открывает шаблон ВОР, заполняет колонку "Количество ГП" нашими объёмами.
    Автодетект колонок по заголовкам, фоллбэк на legacy-значения.
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    detected = _detect_write_columns(ws)
    QTY_PLANNED_COL = detected.get("qty_planned", 9)
    QTY_GP_COL = detected.get("qty_gp", 12)
    PRIMECHANIE_GP_COL = detected.get("gp_note", 19)

    # Если эти колонки есть — заполняем
    stats = {
        "filled": 0,
        "skipped_section_header": 0,
        "no_data": 0,
        "delta_above_30pct": 0,
        "alarms": 0,
        "warns": 0,
    }

    # Найдём первую свободную после 19 колонку для нашего служебного блока
    AUDIT_BASE = 20  # col 20 будет наша audit-метка
    # Заголовки для audit-колонок
    audit_headers = [
        "BIM_qty",       # 20: посчитанное системой
        "BIM_specialist",
        "BIM_confidence",
        "BIM_formula",
        "BIM_status",
        "BIM_delta_pct",
    ]
    for i, h in enumerate(audit_headers):
        cell = ws.cell(1, AUDIT_BASE + i)
        cell.value = h
        cell.font = Font(bold=True, size=10)

    # Идём по строкам шаблона
    for r in range(2, ws.max_row + 1):
        code_v = ws.cell(r, 1).value
        if code_v is None:
            continue
        code_str = str(code_v).strip().rstrip(".")
        if not code_str:
            continue

        allocs = by_position.get(code_str, [])
        if not allocs:
            stats["no_data"] += 1
            continue

        # Если несколько специалистов выдали аллокацию — берём ту с max confidence
        # (или ту что не "delegated_to_*")
        active = [a for a in allocs if not str(a.get("fill_status", "")).startswith("delegated_")]
        if not active:
            active = allocs
        best = max(active, key=lambda a: float(a.get("confidence", 0) or 0))

        qty = best.get("quantity")
        unit = best.get("unit")
        confidence = float(best.get("confidence", 0) or 0)
        formula = best.get("formula", "")
        fill_status = str(best.get("fill_status", ""))
        spec = best.get("_specialist", "?")
        reasoning = best.get("reasoning", "")

        # Заполняем основную колонку qty (только если есть число и confidence > 0)
        if qty is not None and confidence > 0:
            ws.cell(r, QTY_GP_COL).value = qty
            stats["filled"] += 1

        # Comment в col 19 (примечание ГП)
        prim_text = f"[{spec}] conf={confidence:.2f}"
        if reasoning:
            prim_text += f"\n{reasoning[:300]}"
        ws.cell(r, PRIMECHANIE_GP_COL).value = prim_text

        # Audit columns
        ws.cell(r, AUDIT_BASE + 0).value = qty                     # BIM_qty
        ws.cell(r, AUDIT_BASE + 1).value = spec                    # BIM_specialist
        ws.cell(r, AUDIT_BASE + 2).value = round(confidence, 2)    # BIM_confidence
        ws.cell(r, AUDIT_BASE + 3).value = formula                 # BIM_formula
        ws.cell(r, AUDIT_BASE + 4).value = fill_status             # BIM_status

        # Delta vs план
        plan_v = ws.cell(r, QTY_PLANNED_COL).value
        if isinstance(plan_v, (int, float)) and plan_v and qty is not None:
            delta_pct = round(100 * (qty - plan_v) / plan_v, 1)
            ws.cell(r, AUDIT_BASE + 5).value = delta_pct
            if abs(delta_pct) > 30:
                stats["delta_above_30pct"] += 1

        # Подсветка
        fill = confidence_to_fill(confidence, fill_status)
        if fill:
            for c in [QTY_GP_COL, AUDIT_BASE + 0, AUDIT_BASE + 2, AUDIT_BASE + 4]:
                ws.cell(r, c).fill = fill

        if confidence < 0.4:
            stats["alarms"] += 1
        elif confidence < 0.65:
            stats["warns"] += 1

    # Авто-ширина для audit-колонок
    for i in range(len(audit_headers)):
        ws.column_dimensions[get_column_letter(AUDIT_BASE + i)].width = 18

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return stats


def build_audit_workbook(
    output_path: Path,
    specialist_outputs: dict[str, dict],
    overall_summary: dict | None = None,
) -> None:
    """Создаёт отдельную аудит-книгу с разными листами по специалистам."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 1. Сводный лист
    ws = wb.create_sheet("Сводка")
    headers = ["Специалист", "Conf", "Cl-claimed", "Cl-rejected", "Pos-OK", "Pos-Missing", "Concerns", "Summary"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c).value = h
        ws.cell(1, c).font = Font(bold=True)

    r = 2
    for key, data in specialist_outputs.items():
        ph1 = data.get("phase1_filtering", {})
        ph3 = data.get("phase3_allocations", [])
        ph4 = data.get("phase4_gaps", {})
        ws.cell(r, 1).value = key
        ws.cell(r, 2).value = float(data.get("specialist_confidence", 0) or 0)
        ws.cell(r, 3).value = len(ph1.get("claimed", []))
        ws.cell(r, 4).value = len(ph1.get("rejected", []))
        ws.cell(r, 5).value = sum(1 for a in ph3 if a.get("quantity") is not None and float(a.get("confidence", 0) or 0) > 0)
        ws.cell(r, 6).value = len(ph4.get("missing_in_model", []))
        ws.cell(r, 7).value = len(ph4.get("overall_concerns", []))
        ws.cell(r, 8).value = data.get("summary", "")[:200]
        r += 1
    for col_idx, w in enumerate([22, 7, 12, 12, 8, 14, 10, 80], 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # 2. Лист по каждому специалисту
    for key, data in specialist_outputs.items():
        ws = wb.create_sheet(f"S_{key[:25]}")
        # Allocations
        ws.cell(1, 1).value = f"АЛЛОКАЦИИ — {data.get('specialist_name', key)}"
        ws.cell(1, 1).font = Font(bold=True, size=12)
        headers = ["Position", "Qty", "Unit", "Conf", "Status", "Formula", "Reasoning", "Sources"]
        for c, h in enumerate(headers, 1):
            ws.cell(3, c).value = h
            ws.cell(3, c).font = Font(bold=True)
        r = 4
        for alloc in data.get("phase3_allocations", []):
            ws.cell(r, 1).value = alloc.get("position_id")
            ws.cell(r, 2).value = alloc.get("quantity")
            ws.cell(r, 3).value = alloc.get("unit")
            ws.cell(r, 4).value = float(alloc.get("confidence", 0) or 0)
            ws.cell(r, 5).value = alloc.get("fill_status")
            ws.cell(r, 6).value = alloc.get("formula", "")
            ws.cell(r, 7).value = alloc.get("reasoning", "")[:500]
            srcs = alloc.get("source_clusters", [])
            ws.cell(r, 8).value = "\n".join(
                f"{s.get('cluster_id', '')[:50]}  share={s.get('share', 1)}  contrib={s.get('contribution', '?')}"
                for s in srcs
            )
            # Подсветка
            conf = float(alloc.get("confidence", 0) or 0)
            fill = confidence_to_fill(conf, alloc.get("fill_status", ""))
            if fill:
                for c in range(1, 9):
                    ws.cell(r, c).fill = fill
            ws.row_dimensions[r].height = 60
            r += 1

        for col_idx, w in enumerate([12, 10, 6, 6, 24, 50, 80, 60], 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        # Concerns на отдельной секции
        r += 2
        ws.cell(r, 1).value = "Замечания по полноте:"
        ws.cell(r, 1).font = Font(bold=True)
        r += 1
        for issue in data.get("phase2_completeness", []):
            sev = issue.get("severity", "info")
            ws.cell(r, 1).value = sev.upper()
            ws.cell(r, 2).value = issue.get("issue", "")
            if sev == "alarm":
                ws.cell(r, 1).fill = FILL_ALARM
            elif sev == "warn":
                ws.cell(r, 1).fill = FILL_WARN
            r += 1

        # Overall concerns
        r += 1
        ws.cell(r, 1).value = "Общие замечания:"
        ws.cell(r, 1).font = Font(bold=True)
        r += 1
        for c in data.get("phase4_gaps", {}).get("overall_concerns", []):
            ws.cell(r, 1).value = c
            r += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ---------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------
def main():
    import sys, io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

    base = Path(r"C:\Users\kuklev.d.s\PycharmProjects\bim2vor\runs\demo_run")
    template = Path(r"C:\Users\kuklev.d.s\Downloads\Расчет ПЗ_ВГК№5 (ЖК)_Версия 4.xlsx")
    spec_outputs = load_specialist_outputs(base / "specialist_outputs")
    print(f"Прочитано выходов специалистов: {len(spec_outputs)} ({list(spec_outputs.keys())})")

    by_pos = consolidate_allocations(spec_outputs)
    print(f"Аллокаций по позициям: {len(by_pos)}")

    # Filled BoQ
    filled = base / "filled_boq_v2.xlsx"
    stats = fill_boq_template(template, filled, by_pos)
    print(f"\nFilled BoQ: {filled}")
    print(f"  filled: {stats['filled']}")
    print(f"  no_data: {stats['no_data']}")
    print(f"  alarms (conf<0.4): {stats['alarms']}")
    print(f"  warns (0.4-0.65): {stats['warns']}")
    print(f"  delta>30%: {stats['delta_above_30pct']}")

    # Audit
    audit = base / "audit.xlsx"
    build_audit_workbook(audit, spec_outputs)
    print(f"\nAudit: {audit}")


if __name__ == "__main__":
    main()
