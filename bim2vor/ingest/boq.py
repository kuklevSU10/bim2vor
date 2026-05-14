# -*- coding: utf-8 -*-
"""
Чтение шаблона ВОР → нормализованные позиции с иерархией и группировкой по разделам.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterator

import openpyxl


# Колонки шаблона ВОР — автодетект по заголовкам
# Ключевые слова → каноническое имя колонки
HEADER_KEYWORDS = {
    "code":         ["номер позиции", "шифр позиции", "код позиции"],
    "seq_num":      ["№ п/п", "n п/п", "№п/п"],
    "classifier":   ["затрата на строительство", "позиция по классификатору"],
    "element_type": ["тип элемента"],
    "material_type":["тип материала"],
    "name":         ["наименование"],
    "unit":         ["ед. изм", "ед.изм", "единица измерения"],
    "qty_planned":  ["количество заказчика", "количество подрядное", "количество подрядн", "кол-во заказчика"],
    "coef_conv":    ["коэфф. перевода", "коэф. перевода"],
    "coef_consumption": ["коэфф. расхода", "коэф. расхода"],
    "qty_gp":       ["количество гп", "кол-во гп"],
    "currency":     ["валюта"],
    "delivery_type":["тип доставки"],
    "delivery_cost":["стоимость доставки"],
    "price_unit":   ["цена за единицу"],
    "total_cost":   ["итоговая сумма"],
    "kp_link":      ["ссылка на кп"],
    "customer_note":["примечание заказчика"],
    "gp_note":      ["примечание гп"],
}

# Фоллбэк: хардкод номеров колонок для совместимости со старым форматом (ВГК№5)
BOQ_COLUMNS_LEGACY = {
    1: "code", 2: "seq_num", 3: "classifier", 4: "element_type",
    5: "material_type", 6: "name", 7: "unit", 8: "qty_planned",
    9: "coef_conv", 10: "coef_consumption", 11: "qty_gp",
    12: "currency", 13: "delivery_type", 14: "delivery_cost",
    15: "price_unit", 16: "total_cost", 17: "kp_link",
    18: "customer_note", 19: "gp_note",
}


def _detect_columns(ws) -> dict[str, int]:
    """Автодетект: читает строку 1 и ищет ключевые слова → {canonical: col_number (1-based)}."""
    mapping: dict[str, int] = {}
    max_col = min(ws.max_column or 30, 30)
    for c in range(1, max_col + 1):
        v = ws.cell(1, c).value
        if not v:
            continue
        header = str(v).strip().lower()
        for canon, keywords in HEADER_KEYWORDS.items():
            if canon in mapping:
                continue
            for kw in keywords:
                if kw in header:
                    mapping[canon] = c
                    break
    return mapping


@dataclass
class BoQPosition:
    code: str | None                # "1.1.3" или "5.2.1"
    section: int | None             # 5 — верхнеуровневый раздел
    parent_code: str | None         # "1.1"
    depth: int                      # 0 = корень, 1 = раздел, 2 = подраздел...
    seq_num: int | None
    name: str
    unit: str | None
    qty_planned: float | None
    classifier: str | None
    is_section_header: bool         # True если у позиции нет unit и нет qty
    excel_row: int
    raw_row: dict[str, str] = field(default_factory=dict)

    def to_dict(self) -> dict:
        from dataclasses import asdict
        return asdict(self)


def _parse_code(code_raw) -> dict:
    """
    Парсит "1.1.3." / "01.01.01" / "10.2.1." / "325.1" → {section, parent, depth, code}.
    Поддерживает ведущие нули (01 → 1) и произвольную глубину вложенности.
    """
    if not code_raw:
        return {"section": None, "parent": None, "depth": 0, "code": None}
    s = str(code_raw).strip().rstrip(".")
    parts = s.split(".")
    if not all(p.isdigit() for p in parts):
        return {"section": None, "parent": None, "depth": 0, "code": s}
    section = int(parts[0])
    parent = ".".join(parts[:-1]) if len(parts) > 1 else None
    code_normalized = ".".join(parts)
    return {
        "section": section,
        "parent": parent,
        "depth": len(parts),
        "code": code_normalized,
    }


class BoQReader:
    """Читает шаблон ВОР в нормализованную форму. Автодетект колонок по заголовкам."""

    def __init__(self, file_path: Path, sheet_name: str | None = None):
        self.file_path = Path(file_path)
        self.sheet_name = sheet_name
        self._col_map: dict[str, int] | None = None

    def _get_col_map(self, ws) -> dict[str, int]:
        """Определяет маппинг колонок: сначала автодетект, фоллбэк на legacy."""
        if self._col_map is not None:
            return self._col_map
        detected = _detect_columns(ws)
        if "name" in detected and "code" in detected:
            self._col_map = detected
            return self._col_map
        self._col_map = BOQ_COLUMNS_LEGACY
        return self._col_map

    def _cell(self, ws, row: int, col_name: str):
        col_map = self._get_col_map(ws)
        col = col_map.get(col_name)
        if col is None:
            return None
        return ws.cell(row, col).value

    def iter_positions(self) -> Iterator[BoQPosition]:
        wb = openpyxl.load_workbook(self.file_path, data_only=True, read_only=False)
        ws = wb[self.sheet_name] if self.sheet_name else wb.active
        col_map = self._get_col_map(ws)

        for r in range(2, ws.max_row + 1):
            code_v = self._cell(ws, r, "code")
            name_v = self._cell(ws, r, "name")
            if not code_v and not name_v:
                continue
            unit_v = self._cell(ws, r, "unit")
            qty_v = self._cell(ws, r, "qty_planned")
            seq_v = self._cell(ws, r, "seq_num")
            cls_v = self._cell(ws, r, "classifier")

            parsed = _parse_code(code_v)
            qty = qty_v if isinstance(qty_v, (int, float)) else None
            unit = str(unit_v).strip() if unit_v else None

            is_header = parsed["depth"] in (0, 1) or (not unit and qty is None)

            raw_row: dict[str, str] = {}
            for name, col in col_map.items():
                v = ws.cell(r, col).value
                if v is not None:
                    raw_row[name] = str(v) if not isinstance(v, (int, float)) else v

            yield BoQPosition(
                code=parsed["code"],
                section=parsed["section"],
                parent_code=parsed["parent"],
                depth=parsed["depth"],
                seq_num=int(seq_v) if isinstance(seq_v, (int, float)) else None,
                name=str(name_v).strip() if name_v else "",
                unit=unit,
                qty_planned=qty,
                classifier=str(cls_v).strip() if cls_v else None,
                is_section_header=is_header,
                excel_row=r,
                raw_row=raw_row,
            )

    def positions_for_sections(self, sections: list[int]) -> list[BoQPosition]:
        """Возвращает ТОЛЬКО позиции из указанных верхнеуровневых разделов
        (без header'ов уровня 1)."""
        result = []
        for p in self.iter_positions():
            if p.section in sections and not p.is_section_header:
                result.append(p)
        return result

    def all_positions(self) -> list[BoQPosition]:
        return list(self.iter_positions())


def main():
    import sys, io, json
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

    fp = Path(r"C:\Users\kuklev.d.s\Downloads\Расчет ПЗ_ВГК№5 (ЖК)_Версия 4.xlsx")
    reader = BoQReader(fp)

    # Прочитаем всё
    all_pos = reader.all_positions()
    print(f"Всего позиций: {len(all_pos)}")

    # Группировка по разделам
    from collections import defaultdict
    by_section = defaultdict(list)
    for p in all_pos:
        by_section[p.section].append(p)

    print(f"\n=== Разделы (с подсчётом «считаемых» позиций) ===")
    for s in sorted(k for k in by_section.keys() if k):
        pos = by_section[s]
        countable = [p for p in pos if not p.is_section_header]
        with_unit = [p for p in countable if p.unit]
        with_planned = [p for p in countable if p.qty_planned]
        # Достанем имя раздела
        title = next((p.name for p in pos if p.depth == 1 and p.section == s), "?")[:60]
        print(f"  {s:>3}: {len(pos):>3} позиций, {len(countable):>3} считаемых, {len(with_unit):>3} с ед.изм., {len(with_planned):>3} с qty | {title}")

    # Покажем что в разделе 5 (Несущие подземные)
    print(f"\n=== Раздел 5 — Несущие конструкции подземной части ===")
    for p in by_section[5][:30]:
        marker = "▼" if p.is_section_header else "  "
        unit = p.unit or "—"
        qty = p.qty_planned or "—"
        print(f"  {marker} {p.code:>10}  {unit:>5}  {qty!s:>10}  | {p.name[:80]}")


if __name__ == "__main__":
    main()
