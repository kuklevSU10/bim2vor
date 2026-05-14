# -*- coding: utf-8 -*-
"""
Чтение Revit-экспорта (DDC Excel) → нормализованный массив элементов.
Использует openpyxl streaming для больших файлов.
"""
from __future__ import annotations

import json
import re
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterator

import openpyxl

from bim2vor.taxonomy.ost import OstTaxonomy
from bim2vor.parser.family import parse_wall_family


# Канонические колонки которые нам нужны
CORE_COLS = {
    "ID": "element_id",
    "Type Name": "type_name",
    "Category": "category_raw",
    "Family": "family",
    "Level": "level_raw",
    "Volume": "volume",
    "Area": "area",
    "Length": "length",
    "Width": "width",
    "Cost": "cost",
    "Workset": "workset",
    "Description": "description",
    "Comments": "comments",
    "UniqueId": "unique_id",
    "Name": "name",
}


def _strip_type_suffix(header: str) -> str:
    """'Volume : Double' → 'Volume'"""
    if ":" in header:
        return header.split(":")[0].strip()
    return header.strip()


# ---------------------------------------------------------------------
# Level normalizer
# ---------------------------------------------------------------------
LEVEL_PATTERNS = [
    # "6. Этаж", "12 этаж" → number, type=regular
    (re.compile(r"^\s*(?P<n>-?\d+)[\.\s]+(?P<kind>э|Э)таж", re.UNICODE), "regular"),
    (re.compile(r"^\s*Этаж[\s_-]+(?P<kind>повышенный|типовой|технич\w*)", re.UNICODE | re.IGNORECASE), "technical"),
    # "Уровень -1", "Подвал"
    (re.compile(r"^\s*(?P<kind>подвал|цоколь|basement)", re.UNICODE | re.IGNORECASE), "basement"),
    (re.compile(r"^\s*(?P<kind>кровля|крыш|roof)", re.UNICODE | re.IGNORECASE), "roof"),
    (re.compile(r"^\s*(?P<kind>чердак|attic)", re.UNICODE | re.IGNORECASE), "attic"),
    (re.compile(r"^\s*Уровень\s+(?P<n>-?\d+)", re.UNICODE), "regular"),
]


@dataclass
class LevelInfo:
    raw: str
    floor: int | None        # этаж (1-based, отрицательный для подвала)
    zone: str                # regular/basement/roof/technical/attic/unknown


def normalize_level(raw: str | None) -> LevelInfo:
    if not raw:
        return LevelInfo(raw="", floor=None, zone="unknown")
    s = str(raw).strip()
    for pat, default_zone in LEVEL_PATTERNS:
        m = pat.match(s)
        if m:
            n = m.groupdict().get("n")
            zone = default_zone
            kind = m.groupdict().get("kind", "")
            if "технич" in (kind or "").lower():
                zone = "technical"
            elif "повышенный" in (kind or "").lower():
                zone = "technical"
            return LevelInfo(raw=s, floor=int(n) if n else None, zone=zone)
    # Попробуем извлечь число
    m = re.search(r"(-?\d+)", s)
    if m:
        return LevelInfo(raw=s, floor=int(m.group(1)), zone="unknown")
    return LevelInfo(raw=s, floor=None, zone="unknown")


# ---------------------------------------------------------------------
# Element record
# ---------------------------------------------------------------------
@dataclass
class Element:
    element_id: str
    unique_id: str | None
    category_raw: str
    category_canonical: str          # walls/floors/doors/...
    is_physical: bool
    is_excluded: bool
    excluded_reason: str | None
    family: str | None
    type_name: str | None
    level_raw: str | None
    level_floor: int | None
    level_zone: str
    volume_m3: float | None
    area_m2: float | None
    length_m: float | None
    width_m: float | None
    workset: str | None
    description: str | None
    comments: str | None
    cost: float | None
    family_parsed: dict | None = None
    extra: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict:
        from dataclasses import asdict
        return asdict(self)


# ---------------------------------------------------------------------
# Reader
# ---------------------------------------------------------------------
class RevitReader:
    """Стриминговое чтение Revit Excel + нормализация."""

    def __init__(self, file_path: Path, taxonomy: OstTaxonomy | None = None):
        self.file_path = Path(file_path)
        self.taxonomy = taxonomy or OstTaxonomy()
        self._wb: openpyxl.Workbook | None = None
        self._col_map: dict[str, int] = {}    # canonical_name → col_index
        self._sheet_name: str | None = None
        self._headers: list[str] = []

    def _open(self):
        if self._wb is None:
            self._wb = openpyxl.load_workbook(self.file_path, data_only=True, read_only=True)
            self._sheet_name = self._wb.sheetnames[0]
        return self._wb

    def _read_headers(self):
        wb = self._open()
        ws = wb[self._sheet_name]
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        self._headers = [str(h) if h else "" for h in first]
        # ВАЖНО: DDC-экспорт содержит много дубликатов имён колонок
        # ("Family" встречается в шапке многократно для параметров разных типов).
        # Берём ПЕРВОЕ вхождение — это всегда main-параметр элемента.
        for i, h in enumerate(self._headers):
            stripped = _strip_type_suffix(h)
            if stripped in CORE_COLS:
                canon = CORE_COLS[stripped]
                if canon not in self._col_map:
                    self._col_map[canon] = i

    def iter_elements(self) -> Iterator[Element]:
        if not self._col_map:
            self._read_headers()
        wb = self._open()
        ws = wb[self._sheet_name]

        for row in ws.iter_rows(min_row=2, values_only=True):
            cat_raw = self._get(row, "category_raw")
            cat_info = self.taxonomy.classify(str(cat_raw) if cat_raw else None)

            family = self._get(row, "family")
            family = str(family) if family else None
            level_raw = self._get(row, "level_raw")
            level_info = normalize_level(str(level_raw) if level_raw else None)

            family_parsed = None
            if cat_info.canonical == "walls" and family:
                pf = parse_wall_family(family)
                family_parsed = pf.to_dict()

            elem = Element(
                element_id=str(self._get(row, "element_id") or ""),
                unique_id=self._get_str(row, "unique_id"),
                category_raw=cat_info.raw,
                category_canonical=cat_info.canonical,
                is_physical=cat_info.is_physical,
                is_excluded=cat_info.is_excluded,
                excluded_reason=cat_info.excluded_reason,
                family=family,
                type_name=self._get_str(row, "type_name"),
                level_raw=level_info.raw,
                level_floor=level_info.floor,
                level_zone=level_info.zone,
                volume_m3=self._get_float(row, "volume"),
                area_m2=self._get_float(row, "area"),
                length_m=self._get_float(row, "length"),
                width_m=self._get_float(row, "width"),
                workset=self._get_str(row, "workset"),
                description=self._get_str(row, "description"),
                comments=self._get_str(row, "comments"),
                cost=self._get_float(row, "cost"),
                family_parsed=family_parsed,
            )
            yield elem

    def _get(self, row: tuple, name: str) -> Any:
        idx = self._col_map.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    def _get_str(self, row: tuple, name: str) -> str | None:
        v = self._get(row, name)
        if v is None or v == "":
            return None
        return str(v)

    def _get_float(self, row: tuple, name: str) -> float | None:
        v = self._get(row, name)
        if isinstance(v, (int, float)):
            return float(v)
        return None


# ---------------------------------------------------------------------
# CLI: тестируем на реальной выгрузке
# ---------------------------------------------------------------------
def main():
    import sys, io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

    fp = Path(r"C:\Users\kuklev.d.s\Downloads\программа\SKLNK_АР_ПД_К2.1_R25_rvt.xlsx")
    reader = RevitReader(fp)

    stats = Counter()
    physical_with_no_volume = []
    levels_seen = Counter()
    sample_elements = []

    for i, elem in enumerate(reader.iter_elements()):
        stats[elem.category_canonical] += 1
        if elem.is_physical and elem.volume_m3 is None and elem.area_m2 is None:
            physical_with_no_volume.append(elem.category_raw)
        levels_seen[(elem.level_zone, elem.level_floor)] += 1
        if i < 3 and elem.is_physical:
            sample_elements.append(elem)

    print("=== Категории (нормализованные) ===")
    for c, n in stats.most_common(20):
        print(f"  {c:20s}  {n}")

    print(f"\n=== Уровни (zone, floor) — топ 15 ===")
    for (z, f), n in levels_seen.most_common(15):
        print(f"  zone={z:10s} floor={f}  → {n}")

    print(f"\n=== Sample physical elements ===")
    for e in sample_elements:
        print(f"  id={e.element_id} cat={e.category_canonical} family={e.family}")
        print(f"    V={e.volume_m3} A={e.area_m2} level={e.level_floor} zone={e.level_zone}")
        if e.family_parsed:
            print(f"    parsed: {json.dumps(e.family_parsed, ensure_ascii=False)[:200]}")


if __name__ == "__main__":
    main()
