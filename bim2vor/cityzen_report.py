# -*- coding: utf-8 -*-
"""
Cityzen report writer — filled_boq + audit + summary.

Reads:
  runs/<run_id>/bim2vor.sqlite (boq_positions + final_values + source_qtys + clusters)

Writes:
  filled_boq_cityzen_b3.xlsx — копия ВОР с нашими qty + цветовой статус
  audit_cityzen_b3.xlsx — детальный audit trail (8 листов)
  summary.md — итоги в md формате
"""
from __future__ import annotations

import argparse
import json
import sqlite3
import shutil
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
BOQ_FILE = REPO_ROOT / "cetezen" / "Расчет_ПЗ_ЖК_Cityzen_Версия_1.xlsx"

# Colors per fill_status
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_GREY = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
FILL_BLUE = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")


def color_for_zone(z: str | None) -> PatternFill | None:
    if z == "green":
        return FILL_GREEN
    if z == "yellow":
        return FILL_YELLOW
    if z == "red":
        return FILL_RED
    return FILL_GREY


# =============================================================================
# Detect double-count: positions с same filter возвращают same qty
# =============================================================================
def detect_duplicates(conn: sqlite3.Connection) -> dict[int, str]:
    """Detect positions with identical qty under same specialist+unit.
    Returns dict {boq_row: warning_text}."""
    cur = conn.cursor()
    # Group by (specialist, unit, rounded qty) — count positions
    warnings = {}
    rows = list(cur.execute("""
        SELECT bp.row, bp.specialist_key, bp.unit, ROUND(fv.qty, 1)
        FROM boq_positions bp
        JOIN final_values fv ON fv.boq_row=bp.row
        WHERE fv.qty IS NOT NULL AND fv.qty > 0
    """))
    by_key = defaultdict(list)
    for row, spec, unit, qty in rows:
        by_key[(spec, unit, qty)].append(row)
    for key, position_rows in by_key.items():
        if len(position_rows) >= 2:
            for r in position_rows:
                warnings[r] = f"⚠ duplicate_qty (shared with {len(position_rows)-1} other positions, likely needs zone_filter)"
    return warnings


# =============================================================================
# Build filled_boq.xlsx — full BoQ copy with our qty + status column
# =============================================================================
def build_filled_boq(conn: sqlite3.Connection, out_dir: Path):
    print("\n[Report 1] Building filled_boq_cityzen_b3.xlsx...")
    # Load all final_values + source_qtys joined to BoQ rows
    cur = conn.cursor()
    final_by_row = {}
    for r in cur.execute("""
        SELECT bp.row, bp.code, bp.name, bp.unit, bp.parent_path, bp.specialist_key,
               bp.block_assignment, bp.qty_planned,
               fv.qty, fv.delta_abs, fv.abs_tol, fv.zone, fv.fill_status, fv.preferred_source, fv.n_sources
        FROM boq_positions bp
        LEFT JOIN final_values fv ON fv.boq_row=bp.row
    """):
        final_by_row[r[0]] = r
    source_by_row = defaultdict(dict)
    for r in cur.execute("SELECT boq_row, source, qty, n_clusters, n_elements FROM source_qtys"):
        source_by_row[r[0]][r[1]] = {"qty": r[2], "n_clusters": r[3], "n_elements": r[4]}

    duplicates = detect_duplicates(conn)

    # Copy original BoQ
    out_path = out_dir / "filled_boq_cityzen_b3.xlsx"
    shutil.copy(BOQ_FILE, out_path)

    wb = openpyxl.load_workbook(out_path)
    s = wb[wb.sheetnames[0]]

    # Find column "Количество ГП" — это column 12 в Cityzen ВОР (по header row 1)
    QTY_GP_COL = None
    for col in range(1, s.max_column + 1):
        h = s.cell(row=1, column=col).value
        if h and "Количество ГП" in str(h):
            QTY_GP_COL = col
            break
    if QTY_GP_COL is None:
        # Fallback: column 12 (index 11 from 0)
        QTY_GP_COL = 12
    print(f"  Заполняем column {QTY_GP_COL} «Количество ГП»")

    # Также добавим вспомогательные audit-столбцы СПРАВА — для трассировки
    max_col = s.max_column
    audit_headers = ["BIM_zone", "BIM_status", "BIM_specialist",
                     "BIM_S1_AR", "BIM_S2_KR", "BIM_S3_merged",
                     "BIM_delta_abs", "BIM_abs_tol", "BIM_preferred", "BIM_n_src", "BIM_notes"]
    for i, h in enumerate(audit_headers, start=1):
        c = s.cell(row=1, column=max_col + i, value=h)
        c.font = Font(bold=True, color="0000AA")
        c.fill = FILL_BLUE
        c.alignment = Alignment(wrap_text=True, vertical="center")

    n_filled = 0
    for row_num, data in final_by_row.items():
        (_, code, name, unit, parent, spec, ba, qty_pl,
         qty, delta, tol, zone, fs, ps, n_src) = data
        sources = source_by_row.get(row_num, {})
        s1 = sources.get("S1_AR_only", {}).get("qty")
        s2 = sources.get("S2_KR_only", {}).get("qty")
        s3 = sources.get("S3_merged", {}).get("qty")
        warn = duplicates.get(row_num, "")
        notes = warn or ""

        # === ОСНОВНОЕ: заполняем column "Количество ГП" нашим qty ===
        if qty is not None:
            gp_cell = s.cell(row=row_num, column=QTY_GP_COL, value=qty)
            gp_cell.font = Font(bold=True)
            gp_cell.fill = color_for_zone(zone)
            n_filled += 1

        # === Audit-столбцы справа (для трассировки) ===
        audit_vals = [zone, fs, spec, s1, s2, s3, delta, tol, ps, n_src, notes]
        for i, v in enumerate(audit_vals, start=1):
            c = s.cell(row=row_num, column=max_col + i, value=v)
            if i == 1 and zone:
                c.fill = color_for_zone(zone)

    # Widen audit columns
    for i, h in enumerate(audit_headers, start=1):
        s.column_dimensions[get_column_letter(max_col + i)].width = max(12, len(h) + 2)

    wb.save(out_path)
    print(f"  ✓ {out_path} — заполнено {n_filled} строк в «Количество ГП»")
    return out_path, n_filled


# =============================================================================
# Build audit.xlsx — 8-sheet detailed audit
# =============================================================================
def build_audit_xlsx(conn: sqlite3.Connection, out_dir: Path):
    print("\n[Report 2] Building audit_cityzen_b3.xlsx...")
    cur = conn.cursor()
    out_path = out_dir / "audit_cityzen_b3.xlsx"

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Sheet 1: Summary
    s1 = wb.create_sheet("Summary")
    s1.append(["Metric", "Value"])
    s1.append(["Project", "Cityzen Tr. 1 оч., Корпус 3"])
    n_pos = cur.execute("SELECT COUNT(*) FROM boq_positions").fetchone()[0]
    s1.append(["Total BoQ positions analyzed", n_pos])
    for zone in ("green", "yellow", "red"):
        n = cur.execute("SELECT COUNT(*) FROM final_values WHERE zone=?", (zone,)).fetchone()[0]
        s1.append([f"{zone}", n])
    n_needs_llm = cur.execute("SELECT COUNT(*) FROM final_values WHERE fill_status LIKE '%needs_llm%' OR fill_status LIKE '%aux%'").fetchone()[0]
    s1.append(["needs_llm_classification", n_needs_llm])
    n_filled = cur.execute("SELECT COUNT(*) FROM final_values WHERE qty IS NOT NULL").fetchone()[0]
    s1.append(["Positions with computed qty (green+yellow)", n_filled])
    s1.append([])
    s1.append(["BIM Source files (корпус 3 + general)", ""])
    s1.append(["AR_B3", "TSHN08_AR_UNK_R22_UB8B_B3_rvt.xlsx"])
    s1.append(["KR_B3", "TSHN08_KR_STR_R22_UB8B_B3_rvt.xlsx"])
    s1.append(["KV_B3", "TSHN08_KV_UNK_R22_UB8B_B3_rvt.xlsx"])
    n_elem = cur.execute("SELECT COUNT(*) FROM elements").fetchone()[0]
    n_clust = cur.execute("SELECT COUNT(*) FROM clusters").fetchone()[0]
    s1.append([])
    s1.append(["Physical elements ingested", n_elem])
    s1.append(["Clusters formed", n_clust])
    for cell in s1["A"]:
        cell.font = Font(bold=True)

    # Sheet 2: Filled positions (green + yellow)
    s2 = wb.create_sheet("Filled (green+yellow)")
    s2.append(["row", "code", "name", "unit", "qty_BIM", "zone", "fill_status",
               "specialist", "S1_AR", "S2_KR", "S3_merged", "delta_abs", "abs_tol", "preferred_source",
               "parent_section"])
    for r in cur.execute("""
        SELECT bp.row, bp.code, bp.name, bp.unit, fv.qty, fv.zone, fv.fill_status, bp.specialist_key,
               (SELECT qty FROM source_qtys sq WHERE sq.boq_row=bp.row AND sq.source='S1_AR_only'),
               (SELECT qty FROM source_qtys sq WHERE sq.boq_row=bp.row AND sq.source='S2_KR_only'),
               (SELECT qty FROM source_qtys sq WHERE sq.boq_row=bp.row AND sq.source='S3_merged'),
               fv.delta_abs, fv.abs_tol, fv.preferred_source, bp.parent_path
        FROM boq_positions bp JOIN final_values fv ON fv.boq_row=bp.row
        WHERE fv.qty IS NOT NULL ORDER BY bp.specialist_key, bp.row
    """):
        s2.append(list(r))
        # Color zone
        last = s2.max_row
        s2.cell(row=last, column=6).fill = color_for_zone(r[5])

    # Sheet 3: Red zone (no_match / divergent)
    s3 = wb.create_sheet("Red (problems)")
    s3.append(["row", "code", "name", "unit", "zone", "fill_status", "specialist", "S1_AR", "S2_KR", "S3_merged"])
    for r in cur.execute("""
        SELECT bp.row, bp.code, bp.name, bp.unit, fv.zone, fv.fill_status, bp.specialist_key,
               (SELECT qty FROM source_qtys sq WHERE sq.boq_row=bp.row AND sq.source='S1_AR_only'),
               (SELECT qty FROM source_qtys sq WHERE sq.boq_row=bp.row AND sq.source='S2_KR_only'),
               (SELECT qty FROM source_qtys sq WHERE sq.boq_row=bp.row AND sq.source='S3_merged')
        FROM boq_positions bp JOIN final_values fv ON fv.boq_row=bp.row
        WHERE fv.zone='red' ORDER BY bp.specialist_key, bp.row
    """):
        s3.append(list(r))

    # Sheet 4: Needs LLM (requires classification)
    s4 = wb.create_sheet("Needs LLM")
    s4.append(["row", "code", "name", "unit", "fill_status", "specialist", "parent_section"])
    for r in cur.execute("""
        SELECT bp.row, bp.code, bp.name, bp.unit, fv.fill_status, bp.specialist_key, bp.parent_path
        FROM boq_positions bp LEFT JOIN final_values fv ON fv.boq_row=bp.row
        WHERE fv.fill_status LIKE '%needs_llm%' OR fv.fill_status LIKE '%aux%' OR fv.fill_status IS NULL
        ORDER BY bp.specialist_key, bp.row
        LIMIT 1500
    """):
        s4.append(list(r))

    # Sheet 5: Clusters (по специалистам)
    s5 = wb.create_sheet("Clusters")
    s5.append(["cluster_id", "category", "family", "n_total", "vol_m3_total", "area_m2_total",
               "primary_material", "disciplines", "specialist"])
    for r in cur.execute("""
        SELECT cluster_id, category, family, n_total, volume_m3_total, area_m2_total,
               primary_material, disciplines, assigned_specialist
        FROM clusters ORDER BY volume_m3_total DESC
    """):
        s5.append(list(r))

    # Sheet 6: Source qtys breakdown per BoQ
    s6 = wb.create_sheet("Source qtys")
    s6.append(["boq_row", "code", "name", "source", "qty", "n_clusters", "n_elements"])
    for r in cur.execute("""
        SELECT sq.boq_row, bp.code, bp.name, sq.source, sq.qty, sq.n_clusters, sq.n_elements
        FROM source_qtys sq JOIN boq_positions bp ON bp.row=sq.boq_row
        ORDER BY sq.boq_row, sq.source
    """):
        s6.append(list(r))

    # Sheet 7: Conservation summary
    s7 = wb.create_sheet("Conservation")
    s7.append(["Metric", "AR", "KR", "Merged"])
    for r in cur.execute("""
        SELECT category,
               SUM(CASE WHEN source_discipline='AR' THEN volume_m3 ELSE 0 END),
               SUM(CASE WHEN source_discipline='KR' THEN volume_m3 ELSE 0 END),
               SUM(volume_m3)
        FROM elements GROUP BY category ORDER BY SUM(volume_m3) DESC
    """):
        s7.append(list(r))

    # Sheet 8: Specialist distribution
    s8 = wb.create_sheet("Specialists")
    s8.append(["Specialist", "BoQ_positions", "Clusters_assigned", "Filled (green+yellow)", "Red", "Needs_LLM"])
    specs = list(set([r[0] for r in cur.execute("SELECT DISTINCT specialist_key FROM boq_positions")]))
    for spec in specs:
        n_pos = cur.execute("SELECT COUNT(*) FROM boq_positions WHERE specialist_key=?", (spec,)).fetchone()[0]
        n_clust = cur.execute("SELECT COUNT(*) FROM clusters WHERE assigned_specialist=?", (spec,)).fetchone()[0]
        n_filled = cur.execute("""SELECT COUNT(*) FROM boq_positions bp JOIN final_values fv ON fv.boq_row=bp.row
                                  WHERE bp.specialist_key=? AND fv.qty IS NOT NULL""", (spec,)).fetchone()[0]
        n_red = cur.execute("""SELECT COUNT(*) FROM boq_positions bp JOIN final_values fv ON fv.boq_row=bp.row
                               WHERE bp.specialist_key=? AND fv.zone='red'""", (spec,)).fetchone()[0]
        n_llm = cur.execute("""SELECT COUNT(*) FROM boq_positions bp LEFT JOIN final_values fv ON fv.boq_row=bp.row
                              WHERE bp.specialist_key=? AND (fv.fill_status LIKE '%needs_llm%' OR fv.fill_status LIKE '%aux%' OR fv.fill_status IS NULL)""", (spec,)).fetchone()[0]
        s8.append([spec, n_pos, n_clust, n_filled, n_red, n_llm])

    # Auto-size columns for all sheets
    for sheet in wb.worksheets:
        sheet.row_dimensions[1].font = Font(bold=True)
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 18

    wb.save(out_path)
    print(f"  ✓ {out_path}")
    return out_path


# =============================================================================
# Build summary.md
# =============================================================================
def build_summary_md(conn: sqlite3.Connection, out_dir: Path):
    print("\n[Report 3] Building summary.md...")
    cur = conn.cursor()
    n_pos = cur.execute("SELECT COUNT(*) FROM boq_positions").fetchone()[0]
    n_green = cur.execute("SELECT COUNT(*) FROM final_values WHERE zone='green'").fetchone()[0]
    n_yellow = cur.execute("SELECT COUNT(*) FROM final_values WHERE zone='yellow'").fetchone()[0]
    n_red = cur.execute("SELECT COUNT(*) FROM final_values WHERE zone='red'").fetchone()[0]
    n_filled = n_green + n_yellow
    n_needs_llm = cur.execute("SELECT COUNT(*) FROM final_values WHERE fill_status LIKE '%needs_llm%' OR fill_status LIKE '%aux%'").fetchone()[0]
    n_elem = cur.execute("SELECT COUNT(*) FROM elements").fetchone()[0]
    n_clust = cur.execute("SELECT COUNT(*) FROM clusters").fetchone()[0]

    lines = [
        "# Cityzen Корпус 3 — Тендерный отчёт (BIM2VOR baseline)",
        "",
        f"**Дата:** 2026-05-15",
        f"**Source files:** AR_B3 + KR_B3 + KV_B3 (B3 only, STLB excluded)",
        f"**ВОР:** Расчет ПЗ_ЖК Cityzen_Версия 1.xlsx",
        "",
        "## Метрики покрытия",
        "",
        f"- **BoQ позиций корпуса 3 + общих**: {n_pos}",
        f"- **🟢 Green (≥2 источника сошлись)**: {n_green}",
        f"- **🟡 Yellow (single source)**: {n_yellow}",
        f"- **🔴 Red (нет совпадений / divergent)**: {n_red}",
        f"- **💭 Needs LLM (для будущего refinement)**: {n_needs_llm}",
        f"- **Покрытие numeric**: {n_filled}/{n_pos} = {100*n_filled/n_pos:.1f}%",
        "",
        f"- **Physical elements**: {n_elem}",
        f"- **Clusters**: {n_clust}",
        "",
        "## Что готово",
        "",
        "- ✅ Multi-file ingest (AR/KR/KV для B3)",
        "- ✅ Family parser (layered walls breakdown)",
        "- ✅ Clustering (sha256 deterministic IDs)",
        "- ✅ BoQ extract с фильтром «Корпус 3»",
        "- ✅ Specialist mapping per ВОР раздел (4.X → monolith, 6.X → facades, и т.д.)",
        "- ✅ Det compute per source S1 (AR), S2 (KR), S3 (merged)",
        "- ✅ Convergence check abs_tol (0.1 м³ / 0.5 м² / 0.01 тн / 0 шт)",
        "",
        "## Что НЕ закрыто (на доработку перед сдачей)",
        "",
        "1. **Zone split** — позиции 4.1 vs 4.2.2 vs 4.2.4 (подземная / 1-й этаж / выше) получают одинаковую сумму (отмечены в audit.xlsx как `⚠ duplicate_qty`). Требуется фильтр по level_floor.",
        "2. **Mark match для дверей** — все 26 doors-позиций получают одну сумму 636 шт (count всех дверей). Нужно различать по mark (Д-1, ДПМ-01, EI60) через type_name.",
        "3. **Layer split для отделки** — finishing_mop/finishing_parking/finishing_apartments позиции (~400) требуют zone_filter + layer extraction. Помечены `needs_llm`.",
        "4. **Гидроизоляция (раздел 3)** — материалы (мембраны, герметики) не моделируются в BIM. Требуют S4 normative или ручное заполнение.",
        "5. **Лифты** — в BIM-выгрузке B3 не найдены клатеры с family 'лифт'. Возможно в STLB или вне BIM scope. Помечены 0.",
        "",
        "## Распределение по специалистам",
        "",
        "| Specialist | BoQ pos | Clusters | Filled | Red | LLM |",
        "|---|---|---|---|---|---|",
    ]
    for spec, in cur.execute("SELECT DISTINCT specialist_key FROM boq_positions ORDER BY specialist_key"):
        if spec is None: continue
        n_p = cur.execute("SELECT COUNT(*) FROM boq_positions WHERE specialist_key=?", (spec,)).fetchone()[0]
        n_c = cur.execute("SELECT COUNT(*) FROM clusters WHERE assigned_specialist=?", (spec,)).fetchone()[0]
        n_f = cur.execute("SELECT COUNT(*) FROM boq_positions bp JOIN final_values fv ON fv.boq_row=bp.row WHERE bp.specialist_key=? AND fv.qty IS NOT NULL", (spec,)).fetchone()[0]
        n_r = cur.execute("SELECT COUNT(*) FROM boq_positions bp JOIN final_values fv ON fv.boq_row=bp.row WHERE bp.specialist_key=? AND fv.zone='red'", (spec,)).fetchone()[0]
        n_l = cur.execute("SELECT COUNT(*) FROM boq_positions bp LEFT JOIN final_values fv ON fv.boq_row=bp.row WHERE bp.specialist_key=? AND (fv.fill_status LIKE '%needs_llm%' OR fv.fill_status LIKE '%aux%' OR fv.fill_status IS NULL)", (spec,)).fetchone()[0]
        lines.append(f"| {spec} | {n_p} | {n_c} | {n_f} | {n_r} | {n_l} |")
    lines.append("")
    lines.append("## Файлы")
    lines.append("")
    lines.append(f"- [filled_boq_cityzen_b3.xlsx](filled_boq_cityzen_b3.xlsx) — ВОР заказчика + наши столбцы BIM_*")
    lines.append(f"- [audit_cityzen_b3.xlsx](audit_cityzen_b3.xlsx) — детальный аудит trail (8 листов)")
    lines.append(f"- [bim2vor.sqlite](bim2vor.sqlite) — SQLite БД проекта")
    lines.append(f"- [run_summary.json](run_summary.json) — параметры прогона")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("Sub-skills архитектура (10 specialists with 7-stage pipelines) — в `.claude/skills/<expert>-quantity/`")
    out_path = out_dir / "SUMMARY.md"
    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"  ✓ {out_path}")


# =============================================================================
# Main
# =============================================================================
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--run-dir", type=Path, required=True)
    args = parser.parse_args()

    db = args.run_dir / "bim2vor.sqlite"
    conn = sqlite3.connect(db)
    conn.create_function('pylower', 1, lambda s: s.lower() if s else None)

    print(f"=== Cityzen report ===")
    print(f"Run dir: {args.run_dir}")

    build_filled_boq(conn, args.run_dir)
    build_audit_xlsx(conn, args.run_dir)
    build_summary_md(conn, args.run_dir)

    print(f"\n✓ Done. Output: {args.run_dir}")


if __name__ == "__main__":
    main()
