# -*- coding: utf-8 -*-
"""
Cityzen tender runner — корпус 3 + стилобат (опционально).

End-to-end:
  1. Multi-file ingest (AR/KR/KV × B3/STLB)  → elements with source_discipline/source_block
  2. Clustering (canonical_key sha256)
  3. BoQ extract — фильтрация по "Корпус 3" / "Стилобат"
  4. Specialist mapping per BoQ position (через keyword)
  5. Briefing builder per specialist
  6. Sub-skill invocation через Anthropic API (10 параллельных Sonnet)
  7. Det compute: S1_AR_only / S2_KR_only / S3_merged
  8. Convergence check (abs_tol)
  9. Conservation (element_disposition + GATE F)
  10. Report writer: filled_boq_cityzen_b3.xlsx + audit.xlsx + summary.md

Запуск:
  python -m bim2vor.cityzen --include-stlb --target B3 \
    --out runs/cityzen_b3_$(date +%Y%m%d_%H%M%S)
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import hashlib
import re
import sqlite3
from collections import defaultdict, Counter
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from pathlib import Path

# Reuse existing modules
from bim2vor.ingest.revit import RevitReader, Element
from bim2vor.ingest.cluster import cluster_elements, Cluster
from bim2vor.taxonomy.ost import OstTaxonomy
from bim2vor.parser.family import parse_wall_family


REPO_ROOT = Path(__file__).resolve().parents[1]
CETEZEN_DIR = REPO_ROOT / "cetezen"

# Source files for КОРПУС 3 (B1 в naming!) + STLB.
# Naming convention in cetezen folder: B1 = corpus 3 (target), B2 = corpus2, B3 = corpus1, STLB = stylobate.
# User confirmation 2026-05-15: "B1 это корпус 3".
BIM_FILES = {
    "AR_B3":   CETEZEN_DIR / "TSHN08_AR_UNK_R22_UB8B_B1_rvt.xlsx",  # B1 в filename = корпус 3
    "KR_B3":   CETEZEN_DIR / "TSHN08_KR_STR_R22_UB8B_B1_rvt.xlsx",
    "KV_B3":   CETEZEN_DIR / "TSHN08_KV_UNK_R22_UB8B_B1_rvt.xlsx",
    "AR_STLB": CETEZEN_DIR / "TSHN08_AR_UNK_R22_UB8B_STLB_rvt.xlsx",
    "KR_STLB": CETEZEN_DIR / "TSHN08_KR_STR_R22_UB8B_STLB_rvt.xlsx",
    "KV_STLB": CETEZEN_DIR / "TSHN08_KV_UNK_R22_UB8B_STLB_rvt.xlsx",
}
BOQ_FILE = CETEZEN_DIR / "Расчет_ПЗ_ЖК_Cityzen_Версия_1.xlsx"

# Abs tolerances из quality_gates.md
ABS_TOL = {
    "м³":   0.1,
    "м3":   0.1,
    "м²":   0.5,
    "м2":   0.5,
    "тн":   0.01,
    "т":    0.01,
    "шт":   0,
    "компл": 0,
    "пог.м": 0.1,
    "пог.м.": 0.1,
    "м":    0.1,
    "м.п.": 0.1,
    "м.п":  0.1,
    "м/п":  0.1,
    "кг":   1,
}

# Specialist mapping by keywords in BoQ title
SPECIALIST_KEYWORDS = {
    "monolith": [
        # Прямые сигналы бетона
        r"\bбетон\b", r"\bмонолит", r"\bжелезобетон", r"\bЖБ\b",
        # Конструктивные элементы
        r"фундамент", r"подбетонк", r"плит[аы] перекр", r"перекрыти",
        r"\bстен[аы]\b.*B\d+", r"\bстен[аы]\b.*ЖБ",
        r"\bколон", r"\bпилон", r"\bбалк[аи]\b", r"\bростверк",
        r"лестничн[аы]\w*\s+(площадк|марш)", r"\bрампа\b",
        r"парапет.*B\d+", r"капител",
        r"трансферн", r"\bсва[яиио]", r"оголов",
        # Cityzen-specific
        r"ЛПлощадк", r"ЛМарш", r"Перекрыти", r"Покрыти",
        r"ФПлита", r"\bФ_", r"B\d{2,3}\s+F\d+", r"перемычк.*ЖБ",
    ],
    "masonry": [
        r"\bкладк", r"\bблок\b", r"\bкирпич", r"газобет",
        r"силикат", r"ячеист", r"перемычк",
    ],
    "roofing": [
        r"кровл", r"крыш", r"гидроизол.+кровл", r"парапет(?!.*B\d+)",
        r"аэратор", r"воронк.*кровел", r"пароизоляц",
    ],
    "facades": [
        r"фасад", r"утепл", r"изоляция\s+наружн", r"штукатурк.*фасад",
        r"навесн", r"витраж", r"светопрозрач", r"\bокн[ао]\b",
        r"подоконник", r"отлив", r"облицов",
    ],
    "finishing_mop": [
        r"МОП", r"мест.{0,5}общ.{0,5}пользов",
        r"холл", r"коридор.*общ", r"лифтов.{0,3}холл", r"тамбур",
    ],
    "finishing_parking": [
        r"паркинг", r"стоянк[аи]", r"парковк", r"тех\.?\s*помещ",
        r"топпинг", r"ИТП", r"насосная", r"разметк[аи]",
    ],
    "finishing_apartments": [
        r"квартир", r"спальн", r"кухн[яи]", r"санузел", r"ванн",
        r"отделк.*чистов", r"shell\s*[&\-]?\s*core", r"white\s*box",
    ],
    "doors": [
        r"\bдвер[ьия]", r"ворота\b", r"люк", r"калитк",
        r"EI\s*\d+", r"противопожарн.+двер",
    ],
    "metal_stairs": [
        r"металл.+лестниц", r"стальн.+лестниц", r"огражд", r"\bперил",
        r"козыр", r"навес\b", r"входн[аы].+групп", r"балкон.+огражд",
    ],
    "elevators": [
        r"лифт", r"эскалат", r"траволат", r"подъёмник.+МГН",
    ],
}

# OST → primary specialist (fallback)
OST_TO_SPECIALIST = {
    "doors": "doors",
    "windows": "facades",
    "stairs": "metal_stairs",   # default; concrete stairs → monolith via material
    "roofs": "roofing",
    "elevator": "elevators",
    "curtain_panels": "facades",
    "curtain_mullions": "facades",
    "ceilings": "finishing_mop",  # ambient default
    "railings": "metal_stairs",
}

BLOCK_KEYWORDS = ["Корпус 3", "корпус 3", "к.3", "к. 3", "Стилобат", "стилобат"]
TARGET_KEYWORDS_B3 = ["Корпус 3", "корпус 3", "к.3", "к. 3"]
TARGET_KEYWORDS_STLB = ["Стилобат", "стилобат"]


# =============================================================================
# Ingest extension — multi-file with source tracking
# =============================================================================
def ingest_all(files: dict[str, Path]) -> list[dict]:
    """Returns list of element dicts with added 'source_discipline' + 'source_block'."""
    out = []
    tax = OstTaxonomy()
    for source_key, path in files.items():
        if not path.exists():
            print(f"  ! missing: {path.name}")
            continue
        discipline, block = source_key.split("_", 1)  # "AR_B3" → ("AR", "B3")
        reader = RevitReader(path, tax)
        n_total, n_physical = 0, 0
        for e in reader.iter_elements():
            n_total += 1
            if e.is_excluded:
                continue
            d = e.to_dict()
            d["source_discipline"] = discipline
            d["source_block"] = block
            d["source_file"] = path.name
            out.append(d)
            if e.is_physical:
                n_physical += 1
        print(f"  {source_key:10s}  total={n_total:>6}  physical={n_physical:>6}  loaded={sum(1 for x in out if x['source_block']==block and x['source_discipline']==discipline)}")
    return out


# =============================================================================
# Cluster — group across all sources, track discipline split
# =============================================================================
def canonical_key(e: dict) -> str:
    """Канонический ключ для кластера. Игнорирует source — sume identical elements across sources."""
    cat = (e.get("category_canonical") or "").lower().strip()
    fam = (e.get("family") or "").lower().strip()
    typ = (e.get("type_name") or "").lower().strip()
    return f"{cat}|{fam}|{typ}"


def make_cluster_id(canonical: str) -> str:
    return "c_" + hashlib.sha256(canonical.encode("utf-8")).hexdigest()[:16]


@dataclass
class ClusterRec:
    cluster_id: str
    category: str
    family: str | None
    type_name: str | None
    n_elements_total: int = 0
    n_per_source: dict[str, int] = field(default_factory=dict)  # {"AR_B3": n, "KR_B3": n}
    volume_m3_per_source: dict[str, float] = field(default_factory=dict)
    area_m2_per_source: dict[str, float] = field(default_factory=dict)
    length_m_per_source: dict[str, float] = field(default_factory=dict)
    count_per_source: dict[str, int] = field(default_factory=dict)
    level_zones: set[str] = field(default_factory=set)
    level_floors: set[int] = field(default_factory=set)
    primary_material: str | None = None
    family_parsed: dict | None = None
    layers: list[dict] | None = None
    is_underground: bool = False
    zone_marker: str | None = None
    element_ids: list[tuple[str, str]] = field(default_factory=list)  # (source_key, element_id)

    def add(self, e: dict):
        sk = f"{e['source_discipline']}_{e['source_block']}"
        self.n_elements_total += 1
        self.n_per_source[sk] = self.n_per_source.get(sk, 0) + 1
        self.volume_m3_per_source[sk] = self.volume_m3_per_source.get(sk, 0.0) + (e.get("volume_m3") or 0)
        self.area_m2_per_source[sk] = self.area_m2_per_source.get(sk, 0.0) + (e.get("area_m2") or 0)
        self.length_m_per_source[sk] = self.length_m_per_source.get(sk, 0.0) + (e.get("length_m") or 0)
        self.count_per_source[sk] = self.count_per_source.get(sk, 0) + 1
        if e.get("level_zone"):
            self.level_zones.add(e["level_zone"])
        if e.get("level_floor") is not None:
            self.level_floors.add(e["level_floor"])
        if not self.family_parsed and e.get("family_parsed"):
            self.family_parsed = e["family_parsed"]
            if self.family_parsed.get("primary_material"):
                self.primary_material = self.family_parsed["primary_material"]
            if self.family_parsed.get("layers"):
                self.layers = self.family_parsed["layers"]
        self.element_ids.append((sk, e["element_id"]))

    @property
    def volume_m3_total(self) -> float:
        return sum(self.volume_m3_per_source.values())

    @property
    def area_m2_total(self) -> float:
        return sum(self.area_m2_per_source.values())

    @property
    def count_total(self) -> int:
        return self.n_elements_total

    @property
    def disciplines(self) -> list[str]:
        return sorted({sk.split("_")[0] for sk in self.n_per_source.keys()})

    @property
    def blocks(self) -> list[str]:
        return sorted({sk.split("_", 1)[1] for sk in self.n_per_source.keys()})

    def to_brief_dict(self) -> dict:
        """Compact rep for briefing — for LLM."""
        out = {
            "cluster_id": self.cluster_id,
            "category": self.category,
            "family": self.family,
            "type_name": self.type_name,
            "n_elements_total": self.n_elements_total,
            "n_per_source": self.n_per_source,
            "volume_m3_total": round(self.volume_m3_total, 3),
            "area_m2_total": round(self.area_m2_total, 3),
            "disciplines": self.disciplines,
            "blocks": self.blocks,
        }
        if self.primary_material:
            out["primary_material"] = self.primary_material
        if self.layers:
            out["layers"] = [{"material": l.get("material"), "thickness_mm": l.get("thickness_mm")} for l in self.layers]
        if self.level_zones:
            out["level_zones"] = sorted(self.level_zones)
        if self.zone_marker:
            out["zone_marker"] = self.zone_marker
        return out


def cluster_all(elements: list[dict], physical_only: bool = True) -> dict[str, ClusterRec]:
    clusters: dict[str, ClusterRec] = {}
    for e in elements:
        if physical_only and not e.get("is_physical"):
            continue
        ck = canonical_key(e)
        cid = make_cluster_id(ck)
        if cid not in clusters:
            clusters[cid] = ClusterRec(
                cluster_id=cid,
                category=e.get("category_canonical", ""),
                family=e.get("family"),
                type_name=e.get("type_name"),
            )
        clusters[cid].add(e)
    return clusters


# =============================================================================
# BoQ extract — filter for "Корпус 3" + optional "Стилобат"
# =============================================================================
@dataclass
class BoQPos:
    row: int
    code: str            # "1.", "4.1", "103" etc — raw
    name: str
    unit: str | None
    qty_planned: float | None  # for reference, NOT used
    parent_path: list[str] = field(default_factory=list)  # ["4. Возведение", "4.1 Подземная часть"]
    is_section_header: bool = False
    block_assignment: str = "unknown"  # "B3" | "STLB" | "general" | "other_block"
    specialist_key: str | None = None
    cost_code_archetype: str | None = None  # e.g. "concrete_underground"


def extract_boq(boq_path: Path, include_stlb: bool = False) -> list[BoQPos]:
    """Parse Cityzen BoQ, filter for Корпус 3 + optional Стилобат.

    Cityzen BoQ structure:
      - Header rows: code типа "4.1", "4.1.2", "4.1.2.1.2" (dotted), unit=None
      - Position rows: code is integer like "104", "108", unit=м3/шт/м2/...

    Block assignment inherits from parent_path if not in title directly.
    """
    import openpyxl
    wb = openpyxl.load_workbook(boq_path, read_only=True, data_only=True)
    s = wb[wb.sheetnames[0]]
    positions: list[BoQPos] = []
    parent_stack: list[str] = []     # [(code, title)]
    for i, row in enumerate(s.iter_rows(values_only=True), start=1):
        if i == 1:
            continue  # header
        code = row[0]
        seq = row[1]
        title = row[6] if len(row) > 6 else None
        unit = row[7] if len(row) > 7 else None
        qty = row[8] if len(row) > 8 else None
        if not title:
            continue
        title_s = str(title).strip()
        code_s = str(code).strip() if code is not None else ""

        # Header detection — Cityzen specific:
        # Cityzen ВОР использует unit="компл" даже для header rows.
        # Реальный признак: qty is empty/None → header, qty has value → position.
        # Edge: missing unit → also header (top-level Лот roots).
        qty_empty = qty is None or (isinstance(qty, str) and qty.strip() == "")
        is_header = qty_empty
        # depth from code dot count (1 for "1" or "4", 2 for "1.1", 5 for "4.1.2.1.2")
        depth = code_s.rstrip(".").count(".") + 1 if code_s else 1

        # Adjust parent stack first (so position rows inherit current stack)
        if is_header:
            parent_stack = parent_stack[: depth - 1]
            parent_stack.append(f"{code_s} {title_s}")

        # Build joined parent_path for block detection
        parent_path = list(parent_stack)
        combined_text = title_s + " " + " ".join(parent_path)

        # Block assignment (look in title AND parent_path)
        ba = "general"
        if "Корпус 3" in combined_text or "Корпус 3 " in combined_text:
            ba = "B3"
        elif any(k in combined_text for k in ("Корпус 1", "Корпус 4", "Корпус 2")):
            ba = "other_block"
        elif any(k in combined_text for k in TARGET_KEYWORDS_STLB):
            ba = "STLB"

        positions.append(
            BoQPos(
                row=i,
                code=code_s,
                name=title_s,
                unit=str(unit).strip() if unit else None,
                qty_planned=float(qty) if isinstance(qty, (int, float)) else None,
                parent_path=parent_path,
                is_section_header=is_header,
                block_assignment=ba,
            )
        )
    wb.close()

    # Filter:
    filtered = []
    for p in positions:
        if p.is_section_header:
            continue
        if not p.unit:
            continue
        # Skip "1.X" Лот/Подготовительные/МОКАП — это всё под "1. Лот", OOS
        # OK keep them, mark as out_of_scope later
        if p.block_assignment == "B3":
            filtered.append(p)
        elif p.block_assignment == "STLB" and include_stlb:
            filtered.append(p)
        elif p.block_assignment == "general":
            filtered.append(p)
    return filtered


SECTION_TO_SPECIALIST = {
    # 1.X = Подготовка/мобилизация — OOS
    "1":   "OUT_OF_BIM_SCOPE",
    # 2.X = Котлован — OOS
    "2":   "OUT_OF_BIM_SCOPE",
    # 3.X = Гидроизоляция → monolith dopniki
    "3":   "monolith",
    # 4.X = Ж/Б конструкции, металлоконструкции
    "4.1": "monolith",
    "4.2": "monolith",
    "4.3": "metal_stairs",
    "4":   "monolith",
    # 5.X = Подземная/надземная отделка фасадных слоёв или подсекций
    "5":   "facades",       # default — but check parent
    # 6.X = Фасады (светопрозрачные, навесные, декор)
    "6":   "facades",
    # 7.X = Кровля
    "7":   "roofing",
    # 8.1 = Отделка подземной — finishing_parking; 8.2 = надземная — split МОП/квартиры/двери;
    "8.1": "finishing_parking",
    "8.2": "finishing_mop",       # default — но содержит и квартиры
    "8.3": "doors",
    "8":   "finishing_mop",
    # 9.X = Лифты + прочее
    "9.1": "elevators",
    "9":   "elevators",  # fallback
    # 10.X = Инженерные системы — OOS
    "10":  "OUT_OF_BIM_SCOPE",
    # 11.X = Благоустройство — OOS
    "11":  "OUT_OF_BIM_SCOPE",
    # 12.X = Технология подземной автостоянки — partial OOS
    "12":  "finishing_parking",
    # 13.X (if any) = ?
    # 14.X = SHELL & CORE отделка квартир
    "14":  "finishing_apartments",
}


def map_specialist(p: BoQPos) -> str | None:
    """Map BoQ position to specialist via section in parent_path[1]."""
    # Level-2 section code from parent_path[1]
    if len(p.parent_path) >= 2:
        sec2 = p.parent_path[1].strip()
        code = sec2.split(" ", 1)[0].rstrip(".") if sec2 else ""
        # Try exact match first
        if code in SECTION_TO_SPECIALIST:
            spec = SECTION_TO_SPECIALIST[code]
            return None if spec == "OUT_OF_BIM_SCOPE" else spec
        # Try top-level
        top = code.split(".")[0]
        if top in SECTION_TO_SPECIALIST:
            spec = SECTION_TO_SPECIALIST[top]
            return None if spec == "OUT_OF_BIM_SCOPE" else spec

    # Fallback: keyword matching (less reliable)
    title = p.name + " " + " ".join(p.parent_path)
    for spec, patterns in SPECIALIST_KEYWORDS.items():
        for pat in patterns:
            if re.search(pat, title, re.IGNORECASE):
                return spec

    # Try deep-level: position в section 8.2 (Отделка) — split by zone keyword in title
    if len(p.parent_path) >= 2:
        sec2 = p.parent_path[1]
        if "8.2" in sec2 or "Отделка надземной" in sec2:
            full_text = p.name + " " + " ".join(p.parent_path)
            if re.search(r"квартир|спальн|кухн|санузел", full_text, re.I):
                return "finishing_apartments"
            return "finishing_mop"

    return None


# =============================================================================
# Bucket map — cluster → specialist (preliminary)
# =============================================================================
def cluster_to_specialist(c: ClusterRec) -> str | None:
    """Determine which specialist a cluster belongs to (via OST + family signal)."""
    fam_text = f"{c.family or ''} {c.type_name or ''}".lower()

    # 1. Try family signal_words
    for spec, patterns in SPECIALIST_KEYWORDS.items():
        for pat in patterns:
            if re.search(pat, fam_text, re.IGNORECASE):
                return spec

    # 2. Fallback by OST
    if c.category in OST_TO_SPECIALIST:
        spec = OST_TO_SPECIALIST[c.category]
        # Material-based overrides
        if spec == "metal_stairs" and c.primary_material == "concrete":
            return "monolith"
        if spec == "monolith" and c.category == "walls" and c.primary_material in ("block", "brick", "aerated_block"):
            return "masonry"
        return spec

    # 3. Walls: by material
    if c.category == "walls":
        if c.primary_material in ("block", "brick", "aerated_block"):
            return "masonry"
        if c.primary_material == "concrete":
            return "monolith"
        if c.primary_material in ("insulation", "plaster", "finish"):
            return "facades"

    if c.category in ("structural_columns", "structural_framing", "foundation"):
        return "monolith"
    if c.category in ("floors",) and c.primary_material == "concrete":
        return "monolith"

    return None


# =============================================================================
# Storage — SQLite
# =============================================================================
def init_db(db_path: Path) -> sqlite3.Connection:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.executescript("""
    DROP TABLE IF EXISTS elements;
    DROP TABLE IF EXISTS clusters;
    DROP TABLE IF EXISTS boq_positions;
    DROP TABLE IF EXISTS element_disposition;
    DROP TABLE IF EXISTS source_qtys;
    DROP TABLE IF EXISTS final_values;
    DROP TABLE IF EXISTS expert_outputs;
    CREATE TABLE elements (
        element_id TEXT,
        source_discipline TEXT,
        source_block TEXT,
        category TEXT,
        family TEXT,
        type_name TEXT,
        volume_m3 REAL,
        area_m2 REAL,
        length_m REAL,
        level_floor INTEGER,
        level_zone TEXT,
        primary_material TEXT,
        cluster_id TEXT,
        is_physical INTEGER,
        is_excluded INTEGER,
        excluded_reason TEXT,
        PRIMARY KEY (source_discipline, source_block, element_id)
    );
    CREATE INDEX elements_by_cluster ON elements(cluster_id);
    CREATE INDEX elements_by_cat ON elements(category);

    CREATE TABLE clusters (
        cluster_id TEXT PRIMARY KEY,
        category TEXT,
        family TEXT,
        type_name TEXT,
        n_total INTEGER,
        n_per_source TEXT,         -- JSON {"AR_B3": 5, ...}
        volume_m3_total REAL,
        area_m2_total REAL,
        primary_material TEXT,
        layers_json TEXT,
        disciplines TEXT,
        blocks TEXT,
        assigned_specialist TEXT
    );
    CREATE TABLE boq_positions (
        row INTEGER PRIMARY KEY,
        code TEXT,
        name TEXT,
        unit TEXT,
        qty_planned REAL,
        parent_path TEXT,
        block_assignment TEXT,
        specialist_key TEXT
    );
    CREATE INDEX boq_by_specialist ON boq_positions(specialist_key);

    CREATE TABLE element_disposition (
        element_id TEXT,
        source_discipline TEXT,
        source_block TEXT,
        disposition_kind TEXT,     -- allocated_main | allocated_dopnik | out_of_scope_* | excluded_* | unmapped
        boq_row INTEGER,
        specialist TEXT,
        notes TEXT,
        PRIMARY KEY (source_discipline, source_block, element_id)
    );

    CREATE TABLE source_qtys (
        boq_row INTEGER,
        source TEXT,               -- S1_AR_only / S2_KR_only / S3_merged / S4_normative_backcalc
        qty REAL,
        n_clusters INTEGER,
        n_elements INTEGER,
        notes TEXT,
        PRIMARY KEY (boq_row, source)
    );
    CREATE TABLE final_values (
        boq_row INTEGER PRIMARY KEY,
        qty REAL,
        delta_abs REAL,
        abs_tol REAL,
        zone TEXT,                 -- green | yellow | red
        fill_status TEXT,
        preferred_source TEXT,
        n_sources INTEGER
    );
    CREATE TABLE expert_outputs (
        specialist TEXT PRIMARY KEY,
        iteration INTEGER,
        output_json TEXT,
        validation_passed INTEGER,
        n_claims INTEGER,
        n_gaps INTEGER,
        n_out_of_scope INTEGER
    );
    """)
    conn.commit()
    return conn


def store_elements(conn, elements: list[dict], cluster_map: dict):
    """cluster_map: element_key (source_disc, source_block, element_id) → cluster_id"""
    rows = []
    for e in elements:
        if not e.get("is_physical"):
            continue
        key = (e["source_discipline"], e["source_block"], e["element_id"])
        cid = cluster_map.get(key)
        rows.append((
            e["element_id"],
            e["source_discipline"],
            e["source_block"],
            e.get("category_canonical"),
            e.get("family"),
            e.get("type_name"),
            e.get("volume_m3"),
            e.get("area_m2"),
            e.get("length_m"),
            e.get("level_floor"),
            e.get("level_zone"),
            (e.get("family_parsed") or {}).get("primary_material") if e.get("family_parsed") else None,
            cid,
            1 if e.get("is_physical") else 0,
            1 if e.get("is_excluded") else 0,
            e.get("excluded_reason"),
        ))
    conn.executemany(
        "INSERT INTO elements VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        rows,
    )
    conn.commit()
    return len(rows)


def store_clusters(conn, clusters: dict[str, ClusterRec], cluster_specialist: dict[str, str | None]):
    rows = []
    for cid, c in clusters.items():
        rows.append((
            cid, c.category, c.family, c.type_name, c.n_elements_total,
            json.dumps(c.n_per_source, ensure_ascii=False),
            round(c.volume_m3_total, 3), round(c.area_m2_total, 3),
            c.primary_material,
            json.dumps(c.layers, ensure_ascii=False) if c.layers else None,
            ",".join(c.disciplines), ",".join(c.blocks),
            cluster_specialist.get(cid),
        ))
    conn.executemany(
        "INSERT INTO clusters VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
        rows,
    )
    conn.commit()


def store_boq(conn, positions: list[BoQPos]):
    rows = []
    for p in positions:
        rows.append((
            p.row, p.code, p.name, p.unit, p.qty_planned,
            " | ".join(p.parent_path),
            p.block_assignment,
            p.specialist_key,
        ))
    conn.executemany("INSERT INTO boq_positions VALUES (?,?,?,?,?,?,?,?)", rows)
    conn.commit()


# =============================================================================
# Main pipeline runner
# =============================================================================
def main():
    parser = argparse.ArgumentParser(description="Cityzen tender runner (корпус 3)")
    parser.add_argument("--include-stlb", action="store_true", help="Include стилобат source files")
    parser.add_argument("--out", type=Path, default=REPO_ROOT / f"runs/cityzen_b3_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
    parser.add_argument("--phase", choices=["ingest", "all"], default="ingest", help="Запуск только ingest или полный")
    args = parser.parse_args()

    out_dir = args.out
    out_dir.mkdir(parents=True, exist_ok=True)
    print(f"=== Cityzen runner ===")
    print(f"Out: {out_dir}")
    print(f"Include STLB: {args.include_stlb}")

    # Select source files
    files = {k: v for k, v in BIM_FILES.items() if args.include_stlb or "_B3" in k}
    print(f"Source files ({len(files)}):")
    for k, v in files.items():
        print(f"  {k}: {v.name}")

    # PHASE 1 — Ingest
    print("\n[PHASE 1] Ingest...")
    elements = ingest_all(files)
    print(f"  Total elements (non-excluded): {len(elements)}")
    physical = [e for e in elements if e.get("is_physical")]
    print(f"  Physical: {len(physical)}")

    # PHASE 2 — Cluster
    print("\n[PHASE 2] Clustering...")
    clusters = cluster_all(elements, physical_only=True)
    print(f"  Clusters: {len(clusters)}")
    # Build cluster_map for elements
    cluster_map = {}
    for cid, c in clusters.items():
        for sk, eid in c.element_ids:
            disc, block = sk.split("_", 1)
            cluster_map[(disc, block, eid)] = cid

    # PHASE 3 — BoQ extract
    print("\n[PHASE 3] BoQ extract...")
    boq_positions = extract_boq(BOQ_FILE, include_stlb=args.include_stlb)
    print(f"  Positions (filtered to B3{' + STLB' if args.include_stlb else ''}, with unit): {len(boq_positions)}")

    # PHASE 4 — Specialist mapping
    print("\n[PHASE 4] Specialist mapping...")
    for p in boq_positions:
        p.specialist_key = map_specialist(p)
    spec_dist = Counter(p.specialist_key for p in boq_positions)
    for spec, n in spec_dist.most_common():
        print(f"  {str(spec):<25}: {n}")

    # Cluster → specialist
    cluster_specialist = {cid: cluster_to_specialist(c) for cid, c in clusters.items()}
    cluster_spec_dist = Counter(v for v in cluster_specialist.values())
    print(f"  Cluster routing:")
    for spec, n in cluster_spec_dist.most_common():
        print(f"    {str(spec):<25}: {n} clusters")

    # PHASE 5 — Store to SQLite
    print("\n[PHASE 5] Storing to SQLite...")
    db_path = out_dir / "bim2vor.sqlite"
    conn = init_db(db_path)
    n_stored = store_elements(conn, elements, cluster_map)
    print(f"  Stored {n_stored} physical elements")
    store_clusters(conn, clusters, cluster_specialist)
    store_boq(conn, boq_positions)
    print(f"  Stored {len(clusters)} clusters, {len(boq_positions)} positions")

    # Run summary
    summary = {
        "run_id": out_dir.name,
        "started_at": datetime.now(timezone.utc).isoformat(),
        "include_stlb": args.include_stlb,
        "n_source_files": len(files),
        "n_elements_total": len(elements),
        "n_physical": len(physical),
        "n_clusters": len(clusters),
        "n_boq_positions": len(boq_positions),
        "boq_by_specialist": dict(spec_dist),
        "clusters_by_specialist": dict(cluster_spec_dist),
        "db_path": str(db_path.relative_to(REPO_ROOT)),
    }
    (out_dir / "run_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nSummary written: {out_dir / 'run_summary.json'}")
    conn.close()
    return summary


if __name__ == "__main__":
    main()
