# -*- coding: utf-8 -*-
"""
Глубокий анализ Revit-экспорта (DDC):
- какие OST_-категории
- какие параметры реально заполнены (не sparse)
- профиль значений: Volume/Area/Length/Width/Cost
- уникальные семейства, имена типов
- покрытие Level
"""
import sys, io, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from collections import Counter, defaultdict
from pathlib import Path

FP = Path(r'C:\Users\kuklev.d.s\Downloads\программа\SKLNK_АР_ПД_К2.1_R25_rvt.xlsx')
OUT = Path(r'C:\Users\kuklev.d.s\PycharmProjects\bim2vor\data\revit_profile.json')
OUT.parent.mkdir(parents=True, exist_ok=True)

wb = openpyxl.load_workbook(FP, data_only=True, read_only=True)
ws = wb[wb.sheetnames[0]]

# Читаем шапку
headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
headers = [(i, str(h) if h else '') for i, h in enumerate(headers_row)]
print(f"Колонок в шапке: {len(headers)}")

# Канонизируем имена колонок
def canon(h):
    """'Type Name : String' → 'type_name', 'Volume : Double' → 'volume'"""
    if ':' in h:
        h = h.split(':')[0]
    return h.strip().lower().replace(' ', '_').replace('-', '_')

col_map = {i: canon(h) for i, h in headers if h}
key_indices = {
    'id': next((i for i, n in col_map.items() if n == 'id'), None),
    'type_name': next((i for i, n in col_map.items() if n == 'type_name'), None),
    'category': next((i for i, n in col_map.items() if n == 'category'), None),
    'family': next((i for i, n in col_map.items() if n == 'family'), None),
    'level': next((i for i, n in col_map.items() if n == 'level'), None),
    'volume': next((i for i, n in col_map.items() if n == 'volume'), None),
    'area': next((i for i, n in col_map.items() if n == 'area'), None),
    'length': next((i for i, n in col_map.items() if n == 'length'), None),
    'width': next((i for i, n in col_map.items() if n == 'width'), None),
    'cost': next((i for i, n in col_map.items() if n == 'cost'), None),
    'workset': next((i for i, n in col_map.items() if n == 'workset'), None),
    'description': next((i for i, n in col_map.items() if n == 'description'), None),
    'comments': next((i for i, n in col_map.items() if n == 'comments'), None),
}
print(f"\nКлючевые индексы: {key_indices}")

# Профилирование по каждой категории: families/types/level/volume_sum/area_sum/count
cat_stats = defaultdict(lambda: {
    'count': 0, 'volume_sum': 0.0, 'area_sum': 0.0,
    'families': Counter(), 'levels': Counter(), 'workset': Counter(),
    'has_volume': 0, 'has_area': 0, 'has_length': 0,
})
all_columns_filled = Counter()  # Кол-во NOT NULL по каждой колонке

print("\nСчитаем...")
n_rows = 0
for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    n_rows += 1
    if n_rows % 10000 == 0:
        print(f"  обработано: {n_rows}")

    cat = row[key_indices['category']] if key_indices['category'] is not None else None
    if not cat:
        continue

    s = cat_stats[str(cat)]
    s['count'] += 1

    fam = row[key_indices['family']] if key_indices['family'] is not None else None
    lvl = row[key_indices['level']] if key_indices['level'] is not None else None
    vol = row[key_indices['volume']] if key_indices['volume'] is not None else None
    area = row[key_indices['area']] if key_indices['area'] is not None else None
    length = row[key_indices['length']] if key_indices['length'] is not None else None
    ws_ = row[key_indices['workset']] if key_indices['workset'] is not None else None

    if fam: s['families'][str(fam)] += 1
    if lvl: s['levels'][str(lvl)] += 1
    if ws_: s['workset'][str(ws_)] += 1
    if isinstance(vol, (int, float)) and vol > 0:
        s['has_volume'] += 1
        s['volume_sum'] += vol
    if isinstance(area, (int, float)) and area > 0:
        s['has_area'] += 1
        s['area_sum'] += area
    if isinstance(length, (int, float)) and length > 0:
        s['has_length'] += 1

    # Профилируем колонки только для top-категорий чтобы не съесть память
    # ...пропускаем на этом этапе

print(f"\nВсего строк данных: {n_rows}")

# Сохраняем профиль
output = {
    'source_file': str(FP),
    'rows_total': n_rows,
    'columns_total': len(headers),
    'sheets': wb.sheetnames,
    'category_count': len(cat_stats),
    'categories': {}
}

for cat, s in sorted(cat_stats.items(), key=lambda x: -x[1]['count']):
    output['categories'][cat] = {
        'count': s['count'],
        'volume_sum': round(s['volume_sum'], 2),
        'area_sum': round(s['area_sum'], 2),
        'has_volume_pct': round(100 * s['has_volume'] / s['count'], 1) if s['count'] else 0,
        'has_area_pct': round(100 * s['has_area'] / s['count'], 1) if s['count'] else 0,
        'has_length_pct': round(100 * s['has_length'] / s['count'], 1) if s['count'] else 0,
        'top_families': s['families'].most_common(20),
        'top_levels': s['levels'].most_common(15),
        'top_worksets': s['workset'].most_common(10),
    }

OUT.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding='utf-8')
print(f"\nПрофиль записан: {OUT}")

# Краткое резюме
print(f"\n=== САМЫЕ ПОЛЕЗНЫЕ КАТЕГОРИИ (с физическим объёмом) ===")
useful = [(c, d) for c, d in output['categories'].items() if d['volume_sum'] > 0 or d['area_sum'] > 100]
useful.sort(key=lambda x: -x[1]['volume_sum'] - x[1]['area_sum']*0.1)
for cat, d in useful[:20]:
    print(f"  {cat:32s} | n={d['count']:6d} | V={d['volume_sum']:>9.0f} м³ | A={d['area_sum']:>10.0f} м² | hasV={d['has_volume_pct']:.0f}% hasA={d['has_area_pct']:.0f}%")
