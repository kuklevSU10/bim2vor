# -*- coding: utf-8 -*-
"""
Детерминистическая перепроверка: пересчитывает количества ИЗ СЫРОЙ БД
и сравнивает с тем, что выдал LLM-специалист.

Цепочка: claimed cluster_ids → SQL SUM(volume/area) → сравнение с allocation qty.
Ни один LLM не участвует — чистая математика.

Также: read-back из filled_boq.xlsx — проверяет что числа реально попали в нужные ячейки.
"""
from __future__ import annotations

import json
import sqlite3
import sys
import io
from collections import defaultdict
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

REPO = Path(__file__).resolve().parents[1]


def load_json(p: Path) -> dict:
    return json.loads(p.read_text(encoding="utf-8"))


# ====================================================================
# 1. Пересчёт из БД
# ====================================================================
def recompute_from_db(
    db_path: Path,
    run_id: str,
    spec_output: dict,
) -> list[dict]:
    """Для каждой аллокации с source_clusters:
    - берём cluster_ids из source_clusters
    - идём в SQL и считаем SUM(volume_m3), SUM(area_m2), COUNT(*)
    - сравниваем с qty из аллокации
    """
    conn = sqlite3.connect(str(db_path))
    results = []

    for alloc in spec_output.get("phase3_allocations", []):
        pid = alloc.get("position_id", "?")
        qty = alloc.get("quantity")
        unit = str(alloc.get("unit", "")).lower().strip()
        confidence = float(alloc.get("confidence", 0) or 0)
        sources = alloc.get("source_clusters", [])

        if qty is None or not sources:
            continue

        # Для каждого source_cluster — запросить БД
        db_total = 0.0
        db_details = []

        for sc in sources:
            cid = sc.get("cluster_id", "")
            share = float(sc.get("share", 1.0))

            # cluster_id = "category::family::type_name"
            parts = cid.split("::")
            if len(parts) < 3:
                continue
            cat, fam, typ = parts[0], parts[1], parts[2]

            # Query raw elements
            query = """
                SELECT COUNT(*) as cnt,
                       SUM(COALESCE(volume_m3, 0)) as vol,
                       SUM(COALESCE(area_m2, 0)) as area,
                       SUM(COALESCE(length_m, 0)) as len
                FROM elements
                WHERE run_id = ?
                  AND category = ?
                  AND COALESCE(family, '') = ?
                  AND COALESCE(type_name, '') = ?
                  AND is_excluded = 0
                  AND (is_physical = 1 OR category IN ('rooms', 'apartment_type'))
            """
            row = conn.execute(query, (run_id, cat, fam, typ)).fetchone()
            if not row:
                continue

            cnt, vol, area, length = row

            # Какую метрику берём зависит от unit позиции
            if "м3" in unit or "м³" in unit:
                raw_val = vol * share
            elif "м2" in unit or "м²" in unit:
                raw_val = area * share
            elif "шт" in unit:
                raw_val = cnt * share
            elif "пог" in unit:
                raw_val = length * share
            elif "тн" in unit or "т" == unit:
                raw_val = vol * share * 2.5  # бетон ~2.5 т/м³
            else:
                raw_val = None

            if raw_val is not None:
                db_total += raw_val
                db_details.append({
                    "cluster_id": cid[:60],
                    "share": share,
                    "db_raw": round(raw_val, 2),
                    "db_vol": round(vol, 2),
                    "db_area": round(area, 2),
                    "db_count": cnt,
                })

        if db_total > 0 and qty is not None:
            delta = abs(qty - db_total)
            delta_pct = (delta / db_total) * 100 if db_total else 0

            results.append({
                "position_id": pid,
                "llm_qty": round(qty, 2),
                "db_qty": round(db_total, 2),
                "unit": unit,
                "delta_abs": round(delta, 2),
                "delta_pct": round(delta_pct, 1),
                "confidence": confidence,
                "match": delta_pct < 5,
                "n_sources": len(db_details),
            })

    conn.close()
    return results


# ====================================================================
# 2. Read-back из filled_boq.xlsx
# ====================================================================
def readback_filled_boq(
    filled_path: Path,
    by_position: dict[str, list[dict]],
) -> list[dict]:
    """Открывает заполненный ВОР и проверяет что значения действительно записаны."""
    if not filled_path.exists():
        return [{"check": "readback", "severity": "alarm", "message": f"Файл не найден: {filled_path}"}]

    wb = openpyxl.load_workbook(filled_path, read_only=True)
    ws = wb.active

    # Detect columns
    qty_gp_col = None
    code_col = None
    for c in range(1, min(ws.max_column or 30, 30) + 1):
        v = ws.cell(1, c).value
        if not v:
            continue
        h = str(v).strip().lower()
        if "количество гп" in h:
            qty_gp_col = c
        if "номер позиции" in h or "шифр позиции" in h:
            code_col = c

    if qty_gp_col is None:
        return [{"check": "readback", "severity": "alarm", "message": "Не найдена колонка 'Количество ГП'"}]

    results = []
    n_match = 0
    n_mismatch = 0
    n_missing = 0

    for r in range(2, ws.max_row + 1):
        code_v = ws.cell(r, code_col or 1).value
        if code_v is None:
            continue
        code_str = str(code_v).strip().rstrip(".")

        allocs = by_position.get(code_str, [])
        if not allocs:
            continue

        # Expected value
        active = [a for a in allocs if not str(a.get("fill_status", "")).startswith("delegated_")]
        if not active:
            active = allocs
        best = max(active, key=lambda a: float(a.get("confidence", 0) or 0))
        expected_qty = best.get("quantity")
        expected_conf = float(best.get("confidence", 0) or 0)

        if expected_qty is None or expected_conf <= 0:
            continue

        # Read actual cell value
        actual = ws.cell(r, qty_gp_col).value

        if actual is None:
            n_missing += 1
            results.append({
                "position_id": code_str,
                "expected": expected_qty,
                "actual": None,
                "status": "MISSING",
            })
        elif isinstance(actual, (int, float)):
            if abs(actual - expected_qty) < 0.01:
                n_match += 1
            else:
                n_mismatch += 1
                results.append({
                    "position_id": code_str,
                    "expected": expected_qty,
                    "actual": actual,
                    "status": "MISMATCH",
                })
        else:
            n_mismatch += 1

    wb.close()

    summary = {
        "check": "readback",
        "matched": n_match,
        "mismatched": n_mismatch,
        "missing_in_excel": n_missing,
        "severity": "ok" if n_mismatch == 0 and n_missing == 0 else "alarm",
    }
    return [summary] + results


# ====================================================================
# 3. End-to-end trace для одной позиции
# ====================================================================
def trace_position(
    db_path: Path,
    run_id: str,
    spec_output: dict,
    position_id: str,
) -> dict:
    """Полная трассировка: raw elements → cluster → allocation → quantity."""
    conn = sqlite3.connect(str(db_path))
    trace = {"position_id": position_id, "steps": []}

    alloc = None
    for a in spec_output.get("phase3_allocations", []):
        if a.get("position_id", "").strip().rstrip(".") == position_id:
            alloc = a
            break

    if not alloc:
        trace["error"] = f"Аллокация для {position_id} не найдена"
        conn.close()
        return trace

    trace["llm_quantity"] = alloc.get("quantity")
    trace["llm_unit"] = alloc.get("unit")
    trace["llm_confidence"] = alloc.get("confidence")
    trace["llm_reasoning"] = alloc.get("reasoning", "")[:300]

    for sc in alloc.get("source_clusters", []):
        cid = sc.get("cluster_id", "")
        share = float(sc.get("share", 1.0))
        parts = cid.split("::")
        if len(parts) < 3:
            continue

        cat, fam, typ = parts[0], parts[1], parts[2]

        # Raw elements from DB
        rows = conn.execute("""
            SELECT element_id, volume_m3, area_m2, length_m, level_zone, source_model
            FROM elements
            WHERE run_id = ? AND category = ?
              AND COALESCE(family, '') = ? AND COALESCE(type_name, '') = ?
              AND is_excluded = 0 AND is_physical = 1
            LIMIT 5
        """, (run_id, cat, fam, typ)).fetchall()

        # Aggregates
        agg = conn.execute("""
            SELECT COUNT(*), SUM(COALESCE(volume_m3, 0)), SUM(COALESCE(area_m2, 0))
            FROM elements
            WHERE run_id = ? AND category = ?
              AND COALESCE(family, '') = ? AND COALESCE(type_name, '') = ?
              AND is_excluded = 0 AND is_physical = 1
        """, (run_id, cat, fam, typ)).fetchone()

        step = {
            "cluster_id": cid[:80],
            "share": share,
            "db_element_count": agg[0] if agg else 0,
            "db_total_volume_m3": round(agg[1], 2) if agg else 0,
            "db_total_area_m2": round(agg[2], 2) if agg else 0,
            "sample_elements": [
                {
                    "id": r[0][:30] if r[0] else "?",
                    "vol": r[1],
                    "area": r[2],
                    "level": r[4],
                    "model": r[5],
                }
                for r in (rows or [])
            ],
        }
        trace["steps"].append(step)

    conn.close()
    return trace


# ====================================================================
# MAIN
# ====================================================================
def main():
    spec_dir = REPO / "runs" / "event_6_1" / "specialist_outputs"
    briefings_dir = REPO / "runs" / "event_6_1" / "briefings"
    db_path = REPO / "runs" / "bim2vor.db"
    filled_path = REPO / "runs" / "event_6_1" / "filled_boq.xlsx"

    conn = sqlite3.connect(str(db_path))
    run_id = conn.execute("SELECT DISTINCT run_id FROM elements ORDER BY run_id DESC LIMIT 1").fetchone()[0]
    conn.close()

    print("=" * 70)
    print("ДЕТЕРМИНИСТИЧЕСКАЯ ПЕРЕПРОВЕРКА")
    print("=" * 70)
    print(f"DB: {db_path}")
    print(f"Run ID: {run_id}")
    print()

    outputs = {}
    for p in sorted(spec_dir.glob("*.json")):
        data = load_json(p)
        outputs[data.get("specialist", p.stem)] = data

    # 1. DB recompute per specialist
    for spec_key, output in outputs.items():
        print(f"─── {spec_key.upper()}: DB vs LLM ───")
        results = recompute_from_db(db_path, run_id, output)

        n_match = sum(1 for r in results if r["match"])
        n_mismatch = sum(1 for r in results if not r["match"])

        for r in results:
            icon = "✓" if r["match"] else "✗"
            print(f"  {icon} {r['position_id']:15s}  LLM={r['llm_qty']:>10.2f}  DB={r['db_qty']:>10.2f}  "
                  f"delta={r['delta_pct']:>5.1f}%  {r['unit']}")

        print(f"  → Match: {n_match}, Mismatch (>5%): {n_mismatch}")
        print()

    # 2. Read-back from filled Excel
    print("─── READ-BACK: filled_boq.xlsx ───")
    sys.path.insert(0, str(REPO))
    from bim2vor.report.writer import load_specialist_outputs, consolidate_allocations
    all_outputs = load_specialist_outputs(spec_dir)
    by_pos = consolidate_allocations(all_outputs)

    rb_results = readback_filled_boq(filled_path, by_pos)
    for r in rb_results:
        if isinstance(r, dict) and "check" in r:
            print(f"  Matched: {r.get('matched', 0)}, Mismatched: {r.get('mismatched', 0)}, "
                  f"Missing: {r.get('missing_in_excel', 0)}")
        elif isinstance(r, dict) and r.get("status") in ("MISMATCH", "MISSING"):
            print(f"  ! {r['position_id']}: expected={r['expected']}, actual={r['actual']} [{r['status']}]")
    print()

    # 3. End-to-end traces for sample positions
    print("─── END-TO-END TRACES ───")
    sample_positions = [
        ("masonry", "08.01.01"),   # газобетон 200мм — самая большая позиция кладки
        ("masonry", "08.01.06"),   # кирпич полнотелый
        ("monolith", "04.01.03.01"),  # фундаментная плита
        ("monolith", "04.02.02"),     # бетонная подготовка
    ]

    for spec_key, pid in sample_positions:
        output = outputs.get(spec_key, {})
        if not output:
            continue
        trace = trace_position(db_path, run_id, output, pid)
        print(f"\n  [{spec_key}] {pid}:")
        print(f"    LLM qty = {trace.get('llm_quantity')} {trace.get('llm_unit')} "
              f"(conf={trace.get('llm_confidence')})")
        for step in trace.get("steps", []):
            print(f"    Cluster: {step['cluster_id']}")
            print(f"      DB: {step['db_element_count']} elements, "
                  f"V={step['db_total_volume_m3']}м³, A={step['db_total_area_m2']}м²")
            if step.get("sample_elements"):
                for e in step["sample_elements"][:3]:
                    print(f"        elem: vol={e['vol']}, area={e['area']}, "
                          f"level={e['level']}, model={e['model']}")
        if trace.get("error"):
            print(f"    ERROR: {trace['error']}")

    print()
    print("=" * 70)


if __name__ == "__main__":
    main()
