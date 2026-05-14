import duckdb
import pandas as pd
import re
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
import os

DB_PATH = r"C:\Users\kuklev.d.s\bim_warehouse.db"
BOQ_TEMPLATE = r"C:\Users\kuklev.d.s\PycharmProjects\bim2vor\runs\event_6_1\vor_comparison — копия.xlsx"
OUTPUT_BOQ = r"C:\Users\kuklev.d.s\PycharmProjects\bim2vor\runs\event_6_1\vor_comparison — копия.xlsx"
OUTPUT_AUDIT = r"C:\Users\kuklev.d.s\PycharmProjects\bim2vor\results\Event_6_1\Audit_Trail_PRO.xlsx"

def normalize_text(text):
    if not isinstance(text, str): return ""
    text = text.lower()
    # Заменяем латинские на кириллические для унификации
    replacements = {'a':'а', 'b':'в', 'c':'с', 'e':'е', 'h':'н', 'k':'к', 'm':'м', 'o':'о', 'p':'р', 't':'т', 'x':'х', 'y':'у'}
    for eng, rus in replacements.items():
        text = text.replace(eng, rus)
    return text

def extract_grade(text):
    text = normalize_text(text)
    match = re.search(r'в(\d+)', text)
    return match.group(0) if match else ""

def get_section(source_file):
    source = str(source_file).lower()
    if 's1' in source: return 'с1' # кириллица
    if 's2' in source or 's3' in source: return 'с2-3' # кириллица
    return 'общее'

def score_match(boq_name, keywords):
    score = 0
    name = normalize_text(boq_name)
    for kw in keywords:
        kw = normalize_text(kw)
        if kw and kw in name:
            score += 2 if kw.startswith('с1') or kw.startswith('с2') else 1
    return score

def pro_mapping():
    conn = duckdb.connect(DB_PATH)
    
    print("Loading BOQ...")
    df_boq = pd.read_excel(BOQ_TEMPLATE, dtype=str)
    df_boq.columns = [str(c).strip() for c in df_boq.columns]
    
    code_col = next((c for c in df_boq.columns if 'Номер позиции' in c), df_boq.columns[0])
    name_col = next((c for c in df_boq.columns if 'Наименование' in c), df_boq.columns[6])
    
    boq_items = df_boq[[code_col, name_col]].dropna(subset=[code_col]).copy()
    boq_items['code'] = boq_items[code_col].astype(str).str.strip()
    boq_items['name'] = boq_items[name_col].astype(str).str.lower()
    
    print("Querying BIM Warehouse...")
    query = """
    SELECT 
        source_file,
        category,
        type_name,
        mark,
        COUNT(*) as count,
        SUM(volume_m3) as volume_m3,
        SUM(area_m2) as area_m2
    FROM v_expert_analysis
    WHERE type_name IS NOT NULL
    GROUP BY 1, 2, 3, 4
    """
    df_bim = conn.execute(query).df()
    
    audit_records = []
    boq_results = {}
    
    print("Running Semantic Mapping (PRO Scoring Mode)...")
    for idx, row in df_bim.iterrows():
        cat = str(row['category']).upper()
        t_name = str(row['type_name'])
        t_name_norm = normalize_text(t_name)
        vol = float(row['volume_m3'] or 0)
        area = float(row['area_m2'] or 0)
        count = int(row['count'] or 0)
        sec = get_section(row['source_file'])
        
        domain_prefix = None
        keywords = []
        val_used = 0
        unit_used = ""
        reason = ""
        
        # --- МОНОЛИТ ФУНДАМЕНТЫ (05) ---
        if 'фундамент' in t_name_norm or 'плита_ф' in t_name_norm:
            domain_prefix = '05.'
            grade = extract_grade(t_name_norm)
            keywords = [sec, 'фундамент', 'плит', grade]
            val_used = vol
            unit_used = "м3"
            reason = f"Фундамент, секция {sec}, бетон {grade}"
            
        # --- МОНОЛИТ КОНСТРУКЦИИ (06) ---
        elif ('OST_STRUCTURALCOLUMNS' in cat or 'OST_COLUMNS' in cat):
            domain_prefix = '06.01'
            grade = extract_grade(t_name_norm)
            keywords = [sec, 'колонн', 'пилон', grade]
            val_used = vol
            unit_used = "м3"
            reason = f"Колонна, секция {sec}, бетон {grade}"

        elif ('OST_FLOORS' in cat):
            domain_prefix = '06.02'
            grade = extract_grade(t_name_norm)
            keywords = [sec, 'перекрыт', 'плит', grade]
            val_used = vol
            unit_used = "м3"
            reason = f"Перекрытие, секция {sec}, бетон {grade}"

        elif ('OST_WALLS' in cat) and (not ('блок' in t_name_norm or 'кирпич' in t_name_norm or 'газобетон' in t_name_norm or 'фасад' in t_name_norm or 'витраж' in t_name_norm or 'панел' in t_name_norm or 'сн-' in t_name_norm)):
            domain_prefix = '06.03'
            grade = extract_grade(t_name_norm)
            keywords = [sec, 'стен', grade]
            val_used = vol
            unit_used = "м3"
            reason = f"Монолитная стена, секция {sec}, бетон {grade}"
            
        # --- КЛАДКА (07) ---
        elif ('блок' in t_name_norm or 'кирпич' in t_name_norm or 'газобетон' in t_name_norm) and 'OST_WALLS' in cat:
            domain_prefix = '07.'
            mat = 'газобетон' if 'газобетон' in t_name_norm else 'керамич' if 'керам' in t_name_norm else 'блок'
            thickness = re.search(r'\d+', t_name_norm)
            thick = thickness.group(0) if thickness else ""
            keywords = [sec, mat, 'кладк', thick]
            val_used = vol
            unit_used = "м3"
            reason = f"Кладка ({mat}), секция {sec}, толщ {thick}"
            
        # --- ФАСАДЫ (10) ---
        elif ('фасад' in t_name_norm or 'витраж' in t_name_norm or 'панел' in t_name_norm or 'сн-' in t_name_norm) and ('OST_WALLS' in cat or 'OST_CurtainWall' in cat):
            domain_prefix = '10.'
            kw = 'витраж' if 'витраж' in t_name_norm else 'фасад'
            keywords = [sec, kw]
            val_used = area
            unit_used = "м2"
            reason = f"Фасад ({kw}), секция {sec}"
            
        if domain_prefix and val_used > 0:
            candidates = boq_items[boq_items['code'].str.startswith(domain_prefix)].copy()
            if not candidates.empty:
                candidates['score'] = candidates['name'].apply(lambda x: score_match(x, keywords))
                candidates['code_len'] = candidates['code'].str.len()
                best_match = candidates.sort_values(by=['score', 'code_len'], ascending=[False, False]).iloc[0]
                
                # Fallback: даже если счет 0, записываем в базовую категорию (самый короткий код)
                if best_match['score'] == 0:
                    best_match = candidates.sort_values(by=['code_len'], ascending=[True]).iloc[0]
                
                mapped_code = best_match['code']
                boq_results[mapped_code] = boq_results.get(mapped_code, 0) + val_used
                
                audit_records.append({
                    "BOQ_Code": mapped_code,
                    "BOQ_Name": best_match['name'],
                    "BIM_Category": cat,
                    "BIM_Type_Name": t_name,
                    "Section": sec,
                    "Count": count,
                    "Value_Used": val_used,
                    "Unit": unit_used,
                    "Reason": reason,
                    "Source_File": row['source_file'],
                    "Match_Score": best_match['score']
                })

    print("Generating Audit Trail...")
    os.makedirs(os.path.dirname(OUTPUT_AUDIT), exist_ok=True)
    df_audit = pd.DataFrame(audit_records)
    if not df_audit.empty:
        df_audit = df_audit.sort_values(by=["BOQ_Code", "Match_Score"], ascending=[True, False])
    df_audit.to_excel(OUTPUT_AUDIT, index=False)
    
    print(f"Filling Excel {BOQ_TEMPLATE}...")
    wb = openpyxl.load_workbook(BOQ_TEMPLATE)
    ws = wb.active
    
    code_col_idx = None
    gp_col_idx = None
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(row=1, column=col).value)
        if 'Номер позиции' in val: code_col_idx = col
        if 'Количество ГП' in val: gp_col_idx = col
    
    if not gp_col_idx: gp_col_idx = 12
    if not code_col_idx: code_col_idx = 1
    
    fill_color = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    
    filled = 0
    for row in range(2, ws.max_row + 1):
        code_val = ws.cell(row=row, column=code_col_idx).value
        if code_val is None: continue
        code = str(code_val).strip()
        
        if code in boq_results:
            cell = ws.cell(row=row, column=gp_col_idx)
            cell.value = round(boq_results[code], 2)
            cell.fill = fill_color
            
            types_used = df_audit[df_audit['BOQ_Code']==code]['BIM_Type_Name'].unique()
            short_types = ", ".join(types_used[:3]) + ("..." if len(types_used) > 3 else "")
            cell.comment = Comment(f"PRO Mapping\nИсточники ({len(types_used)} типов):\n{short_types}", "BIM2VOR PRO")
            filled += 1
            
    wb.save(OUTPUT_BOQ)
    print(f"PRO Execution Complete. Filled {filled} rows.\nAudit saved to {OUTPUT_AUDIT}\nFinal BOQ saved to {OUTPUT_BOQ}")

if __name__ == "__main__":
    pro_mapping()
