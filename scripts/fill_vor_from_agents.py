# -*- coding: utf-8 -*-
"""
Fill VOR Excel template from agent results.

Reads:
  - runs/event_6_1/agent_masonry_result.json   (section 08)
  - runs/event_6_1/agent_facades_result.json   (section 06)
  - runs/event_6_1/specialist_outputs/monolith.json (sections 04-05, old format)

Writes:
  - runs/event_6_1/vor_filled.xlsx  (copy of VOR with col 12 filled)
"""
import json
import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

PROJ = Path(__file__).resolve().parents[1]
RUNS = PROJ / "runs" / "event_6_1"

VOR_TEMPLATE = Path(r"C:\Users\kuklev.d.s\Downloads\Расчет ПЗ_Событие 6.1_Версия 2.xlsx")
VOR_OUTPUT = RUNS / "vor_filled.xlsx"

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


def load_agent_results() -> dict[str, dict]:
    """Load all agent/specialist results into unified {position_id: {qty, unit, confidence, reasoning, source}}."""
    positions = {}

    # 1. Agent masonry (section 08)
    masonry_path = RUNS / "agent_masonry_result.json"
    if masonry_path.exists():
        data = json.loads(masonry_path.read_text(encoding="utf-8"))
        for pos_id, info in data.get("positions", {}).items():
            if "_combined" in pos_id or "_to_" in pos_id:
                continue
            qty = info.get("qty")
            if qty is None:
                continue
            positions[pos_id] = {
                "qty": qty,
                "unit": info.get("unit", ""),
                "confidence": info.get("confidence", 0),
                "reasoning": info.get("reasoning", ""),
                "source": "agent:masonry",
            }

    # 2. Agent facades (section 06)
    facades_path = RUNS / "agent_facades_result.json"
    if facades_path.exists():
        data = json.loads(facades_path.read_text(encoding="utf-8"))
        for pos_id, info in data.get("positions", {}).items():
            if "_combined" in pos_id or "_to_" in pos_id:
                continue
            qty = info.get("qty")
            if qty is None:
                continue
            positions[pos_id] = {
                "qty": qty,
                "unit": info.get("unit", ""),
                "confidence": info.get("confidence", 0),
                "reasoning": info.get("reasoning", ""),
                "source": "agent:facades",
            }

    # 3. Specialist monolith (sections 04-05, old format)
    monolith_path = RUNS / "specialist_outputs" / "monolith.json"
    if monolith_path.exists():
        data = json.loads(monolith_path.read_text(encoding="utf-8"))
        for alloc in data.get("phase3_allocations", []):
            pos_id = alloc.get("position_id", "")
            qty = alloc.get("quantity")
            if qty is None:
                continue
            conf = alloc.get("confidence", 0)
            reasoning = alloc.get("reasoning", "")

            # Skip copy-from-VOR detections: if reasoning says "Принимаю ВОР" and conf < 0.4
            is_vor_copy = "Принимаю ВОР" in reasoning or "ВОР =" in (alloc.get("formula", "") or "")

            positions[pos_id] = {
                "qty": qty,
                "unit": alloc.get("unit", ""),
                "confidence": conf,
                "reasoning": reasoning,
                "source": "specialist:monolith",
                "is_vor_copy": is_vor_copy,
            }

    return positions


def fill_vor(positions: dict[str, dict]):
    """Copy VOR template and fill column 12 with agent quantities."""
    shutil.copy2(str(VOR_TEMPLATE), str(VOR_OUTPUT))
    wb = openpyxl.load_workbook(str(VOR_OUTPUT))
    ws = wb.active

    filled = 0
    skipped_zero = 0
    skipped_low_conf = 0
    skipped_copy = 0
    not_found = []

    position_rows = {}
    for row in ws.iter_rows(min_row=2, max_col=12):
        cell_a = row[0]
        pos_id = str(cell_a.value).strip() if cell_a.value else ""
        if pos_id:
            position_rows[pos_id] = row

    for pos_id, info in sorted(positions.items()):
        qty = info["qty"]
        conf = info["confidence"]
        source = info["source"]
        is_copy = info.get("is_vor_copy", False)

        # Skip qty=0 with low confidence (not modeled)
        if qty == 0 and conf < 0.3:
            skipped_zero += 1
            continue

        # Skip VOR copies with low confidence from monolith specialist
        if is_copy and conf < 0.4:
            skipped_copy += 1
            continue

        if pos_id not in position_rows:
            not_found.append(pos_id)
            continue

        row = position_rows[pos_id]
        cell_qty = row[11]  # col 12 (0-indexed: 11)
        cell_cust = row[8]  # col 9 (0-indexed: 8)

        # Copy detection: if BIM qty matches customer qty within 0.1%
        cust_val = cell_cust.value
        if cust_val and isinstance(cust_val, (int, float)) and cust_val > 0:
            if qty > 0 and abs(qty - cust_val) / cust_val < 0.001:
                if is_copy:
                    skipped_copy += 1
                    continue

        cell_qty.value = round(qty, 2) if qty != 0 else 0

        # Color by confidence
        if conf >= 0.65:
            cell_qty.fill = GREEN
        elif conf >= 0.4:
            cell_qty.fill = YELLOW
        else:
            cell_qty.fill = RED

        cell_qty.number_format = '#,##0.00'
        filled += 1

    wb.save(str(VOR_OUTPUT))
    wb.close()

    print(f"\n{'='*60}")
    print(f"VOR FILL RESULTS")
    print(f"{'='*60}")
    print(f"Filled positions:     {filled}")
    print(f"Skipped (qty=0/low):  {skipped_zero}")
    print(f"Skipped (VOR copy):   {skipped_copy}")
    print(f"Skipped (low conf):   {skipped_low_conf}")
    print(f"Not found in VOR:     {len(not_found)}")
    if not_found:
        print(f"  Missing IDs: {not_found}")
    print(f"\nOutput: {VOR_OUTPUT}")


def print_comparison(positions: dict[str, dict]):
    """Print comparison table: our qty vs customer qty."""
    wb = openpyxl.load_workbook(str(VOR_TEMPLATE), read_only=True, data_only=True)
    ws = wb.active

    vor_data = {}
    for row in ws.iter_rows(min_row=2, max_col=12, values_only=True):
        pos_id = str(row[0]).strip() if row[0] else ""
        name = str(row[6])[:50] if row[6] else ""
        unit = str(row[7]) if row[7] else ""
        cust_qty = row[8]
        if pos_id and cust_qty is not None:
            vor_data[pos_id] = {"name": name, "unit": unit, "cust_qty": cust_qty}
    wb.close()

    print(f"\n{'='*120}")
    print(f"{'Поз.':<16} {'Наименование':<45} {'Ед.':<6} {'Заказчик':>12} {'BIM':>12} {'Δ%':>8} {'Conf':>6} {'Источник'}")
    print(f"{'-'*120}")

    sections = {}
    for pos_id, info in sorted(positions.items()):
        qty = info["qty"]
        conf = info["confidence"]
        source = info["source"]
        is_copy = info.get("is_vor_copy", False)

        if qty == 0 and conf < 0.3:
            continue
        if is_copy and conf < 0.4:
            continue

        vor = vor_data.get(pos_id, {})
        cust = vor.get("cust_qty", 0) or 0
        name = vor.get("name", "")[:45]
        unit = vor.get("unit", "")

        if cust and cust > 0:
            delta = (qty - cust) / cust * 100
            delta_str = f"{delta:+.1f}%"
        else:
            delta_str = "n/a"

        sec = pos_id.split('.')[0]
        if sec not in sections:
            sections[sec] = []
            print()
        sections[sec].append(pos_id)

        marker = ""
        if is_copy:
            marker = " [VOR]"
        if conf < 0.4:
            marker += " [LOW]"

        print(f"{pos_id:<16} {name:<45} {unit:<6} {cust:>12.2f} {qty:>12.2f} {delta_str:>8} {conf:>6.2f} {source}{marker}")

    print(f"\n{'='*120}")


if __name__ == "__main__":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    print("Loading agent results...")
    positions = load_agent_results()
    print(f"Loaded {len(positions)} positions from agents/specialists")

    print("\nComparison table:")
    print_comparison(positions)

    print("\nFilling VOR...")
    fill_vor(positions)
