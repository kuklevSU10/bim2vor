import json
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os

# Настройки
RAW_DATA_PATH = "results/Event_6_1/final_raw_data.json"
BOQ_TEMPLATE = r"C:\Users\kuklev.d.s\PycharmProjects\bim2vor\runs\demo_run\Расчет ПЗ_Событие 6.1_Версия 4.xlsx"
OUTPUT_FILE = "results/Event_6_1/Filled_BOQ_Monolith_Masonry_Facades.xlsx"

def run_mapping():
    print("Loading data...")
    with open(RAW_DATA_PATH, "r", encoding="utf-8") as f:
        raw_data = json.load(f)
    
    mapping_results = {}

    # 1. МОНОЛИТ (Секции 5 и 6)
    print("Mapping Monolith...")
    for item in raw_data['monolith']:
        t = str(item.get('type_name', '')).lower()
        vol = item.get('total_volume', 0)
        if not vol: continue
        
        # Интеллектуальный маппинг по ключевым словам
        if 'фундамент' in t or 'плита_ф' in t:
            code = "05." # Корневой раздел фундаментов
            mapping_results[code] = mapping_results.get(code, 0) + vol
        elif 'колонн' in t:
            code = "06.01" # Колонны
            mapping_results[code] = mapping_results.get(code, 0) + vol
        elif 'перекрыт' in t or 'пер-' in t:
            code = "06.02" # Перекрытия
            mapping_results[code] = mapping_results.get(code, 0) + vol
        elif 'стен' in t and ('бетон' in t or 'кц_' in t):
            code = "06.03" # Монолитные стены
            mapping_results[code] = mapping_results.get(code, 0) + vol

    # 2. КЛАДКА (Секция 7)
    print("Mapping Masonry...")
    for item in raw_data['monolith']: # Стены часто в этом списке
        t = str(item.get('type_name', '')).lower()
        vol = item.get('total_volume', 0)
        if 'блок' in t or 'кирпич' in t or 'газобетон' in t:
            code = "07." # Раздел кладки
            mapping_results[code] = mapping_results.get(code, 0) + vol

    # 3. ФАСАДЫ (Секция 10)
    print("Mapping Facades...")
    for item in raw_data['facades']:
        t = str(item.get('type_name', '')).lower()
        area = item.get('total_area', 0)
        if not area: continue
        
        if 'витраж' in t or 'фасад' in t or 'сн-' in t or 'панел' in t:
            code = "10." # Раздел фасадов
            mapping_results[code] = mapping_results.get(code, 0) + area

    print(f"Opening template {BOQ_TEMPLATE}...")
    wb = openpyxl.load_workbook(BOQ_TEMPLATE)
    ws = wb.active

    # Динамический поиск колонок
    gp_col = None
    code_col = None
    
    # Читаем первую строку для заголовков
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(row=1, column=col).value)
        if 'Номер позиции' in val: code_col = col
        if 'Количество ГП' in val: gp_col = col
    
    if not gp_col: gp_col = 12 # fallback
    if not code_col: code_col = 1 # fallback

    fill_color = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    print(f"Filling Excel (Code Col: {code_col}, GP Col: {gp_col})...")
    filled_count = 0
    for row in range(2, ws.max_row + 1):
        code_val = ws.cell(row=row, column=code_col).value
        if code_val is None: continue
        code = str(code_val).strip()
        
        for map_code, val in mapping_results.items():
            if code.startswith(map_code):
                ws.cell(row=row, column=gp_col).value = round(val, 2)
                ws.cell(row=row, column=gp_col).fill = fill_color
                filled_count += 1
                break

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"Successfully filled {filled_count} rows. Result saved to {OUTPUT_FILE}")

if __name__ == "__main__":
    run_mapping()
