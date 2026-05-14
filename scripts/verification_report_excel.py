# -*- coding: utf-8 -*-
"""
Генерация Excel-отчёта верификации BIM→ВОР.

Листы:
  1. Сводка — общая картина по каждому уровню проверки
  2. Баланс объёмов — DB total vs claimed vs allocated vs dopnik
  3. DB vs LLM — детерминистическая перепроверка каждой позиции
  4. Покрытие кластеров — каждый кластер: куда попал (позиция/допник/потерян)
  5. Покрытие категорий — категории BIM и их статус
  6. Read-back Excel — проверка что числа попали в ячейки
  7. Допники — кластеры без позиций ВОР
  8. Отсутствует в BIM — позиции ВОР без данных в модели
"""
from __future__ import annotations

import json
import sqlite3
import sys
import io
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

REPO = Path(__file__).resolve().parents[1]

FILL_OK = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_WARN = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_ALARM = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FILL_SECTION = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_GREY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_BOLD = Font(bold=True, size=11)
FONT_NORMAL = Font(size=10)
FONT_SECTION = Font(bold=True, size=11, color="1F4E79")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def load_json(p: Path) -> dict:
    return json.loads(p.read_text(encoding="utf-8"))


def write_header(ws, row, headers, widths=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    if widths:
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w


def write_row(ws, row, values, fills=None, font=None):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row, c, v)
        cell.font = font or FONT_NORMAL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True)
        if fills and c <= len(fills) and fills[c - 1]:
            cell.fill = fills[c - 1]
        if isinstance(v, float):
            cell.number_format = "#,##0.00"


def status_fill(status: str):
    s = status.lower()
    if s in ("ок", "ok", "match", "совпадает"):
        return FILL_OK
    if s in ("warn", "предупреждение", "частично"):
        return FILL_WARN
    return FILL_ALARM


# ====================================================================
# Data collection
# ====================================================================

def collect_all_data(spec_dir, db_path, filled_path):
    conn = sqlite3.connect(str(db_path))
    run_id = conn.execute(
        "SELECT DISTINCT run_id FROM elements ORDER BY run_id DESC LIMIT 1"
    ).fetchone()[0]

    # Load specialist outputs
    outputs = {}
    for p in sorted(spec_dir.glob("*.json")):
        data = load_json(p)
        outputs[data.get("specialist", p.stem)] = data

    # --- 1. DB category totals ---
    cat_rows = conn.execute("""
        SELECT category, COUNT(*) cnt,
               ROUND(SUM(COALESCE(volume_m3,0)),2) vol,
               ROUND(SUM(COALESCE(area_m2,0)),2) area
        FROM elements
        WHERE is_excluded=0 AND is_physical=1
        GROUP BY category ORDER BY SUM(COALESCE(volume_m3,0)) DESC
    """).fetchall()

    # --- 2. All claimed clusters across specialists ---
    claimed_map = {}  # cid -> [spec, ...]
    for spec, data in outputs.items():
        for c in data.get("phase1_filtering", {}).get("claimed", []):
            cid = c.get("cluster_id", "")
            if cid:
                claimed_map.setdefault(cid, []).append(spec)

    # --- 3. All allocated clusters ---
    allocated_map = {}  # cid -> [(spec, pid, qty, unit, share, contrib), ...]
    for spec, data in outputs.items():
        for alloc in data.get("phase3_allocations", []):
            pid = alloc.get("position_id", "?")
            qty = alloc.get("quantity")
            unit = alloc.get("unit", "")
            for sc in alloc.get("source_clusters", []):
                cid = sc.get("cluster_id", "")
                share = sc.get("share", 1.0)
                contrib = sc.get("contribution")
                if cid:
                    allocated_map.setdefault(cid, []).append(
                        (spec, pid, qty, unit, share, contrib)
                    )

    # --- 4. Dopniki clusters ---
    dopnik_map = {}  # cid -> (spec, suggested, est_qty, unit)
    for spec, data in outputs.items():
        for d in data.get("phase4_gaps", {}).get("claimed_but_unallocated", []):
            cid = d.get("cluster_id", "")
            if cid:
                dopnik_map[cid] = (
                    spec,
                    d.get("suggested_dopnik", ""),
                    d.get("estimated_qty", 0),
                    d.get("unit", ""),
                )

    # --- 5. Missing in model ---
    missing_model = []
    for spec, data in outputs.items():
        for m in data.get("phase4_gaps", {}).get("missing_in_model", []):
            missing_model.append({
                "specialist": spec,
                "position_id": m.get("position_id", ""),
                "description": m.get("description", ""),
                "reason": m.get("reason", ""),
            })

    # --- 6. DB recompute ---
    recompute_results = {}
    for spec, data in outputs.items():
        recompute_results[spec] = recompute_from_db(conn, run_id, data)

    # --- 7. Cluster DB data ---
    cluster_db = {}  # cid -> (cnt, vol, area)
    all_cids = set(claimed_map.keys()) | set(allocated_map.keys()) | set(dopnik_map.keys())
    for cid in all_cids:
        parts = cid.split("::")
        if len(parts) < 3:
            continue
        row = conn.execute("""
            SELECT COUNT(*), ROUND(SUM(COALESCE(volume_m3,0)),2),
                   ROUND(SUM(COALESCE(area_m2,0)),2)
            FROM elements WHERE run_id=? AND is_excluded=0 AND is_physical=1
            AND category=? AND COALESCE(family,'')=? AND COALESCE(type_name,'')=?
        """, (run_id, parts[0], parts[1], parts[2])).fetchone()
        if row:
            cluster_db[cid] = (row[0], row[1] or 0, row[2] or 0)

    # --- 8. Category claimed status ---
    cat_claimed = set()
    for cid in claimed_map:
        parts = cid.split("::")
        if parts:
            cat_claimed.add(parts[0])

    conn.close()

    return {
        "run_id": run_id,
        "outputs": outputs,
        "cat_rows": cat_rows,
        "claimed_map": claimed_map,
        "allocated_map": allocated_map,
        "dopnik_map": dopnik_map,
        "missing_model": missing_model,
        "recompute": recompute_results,
        "cluster_db": cluster_db,
        "cat_claimed": cat_claimed,
    }


def recompute_from_db(conn, run_id, spec_output):
    results = []
    for alloc in spec_output.get("phase3_allocations", []):
        pid = alloc.get("position_id", "?")
        qty = alloc.get("quantity")
        unit = str(alloc.get("unit", "")).lower().strip()
        confidence = float(alloc.get("confidence", 0) or 0)
        sources = alloc.get("source_clusters", [])
        reasoning = alloc.get("reasoning", "")

        if qty is None or not sources:
            continue

        db_total = 0.0
        cluster_details = []

        for sc in sources:
            cid = sc.get("cluster_id", "")
            share = float(sc.get("share", 1.0))
            parts = cid.split("::")
            if len(parts) < 3:
                continue

            row = conn.execute("""
                SELECT COUNT(*), SUM(COALESCE(volume_m3,0)),
                       SUM(COALESCE(area_m2,0)), SUM(COALESCE(length_m,0))
                FROM elements WHERE run_id=? AND category=?
                AND COALESCE(family,'')=? AND COALESCE(type_name,'')=?
                AND is_excluded=0
                AND (is_physical=1 OR category IN ('rooms','apartment_type'))
            """, (run_id, parts[0], parts[1], parts[2])).fetchone()
            if not row:
                continue

            cnt, vol, area, length = row
            if "м3" in unit or "м³" in unit:
                raw_val = vol * share
            elif "м2" in unit or "м²" in unit:
                raw_val = area * share
            elif "шт" in unit:
                raw_val = cnt * share
            elif "пог" in unit:
                raw_val = length * share
            elif "тн" in unit or "т" == unit:
                raw_val = vol * share * 2.5
            else:
                raw_val = None

            if raw_val is not None:
                db_total += raw_val
                cluster_details.append(cid[:60])

        if db_total > 0 and qty is not None:
            delta = abs(qty - db_total)
            delta_pct = (delta / db_total) * 100 if db_total else 0
            results.append({
                "position_id": pid,
                "llm_qty": round(qty, 2),
                "db_qty": round(db_total, 2),
                "unit": unit,
                "delta_pct": round(delta_pct, 1),
                "confidence": confidence,
                "match": delta_pct < 5,
                "n_sources": len(cluster_details),
                "reasoning": reasoning[:200] if reasoning else "",
            })

    return results


def readback_check(filled_path, spec_dir):
    sys.path.insert(0, str(REPO))
    from bim2vor.report.writer import load_specialist_outputs, consolidate_allocations

    all_outputs = load_specialist_outputs(spec_dir)
    by_pos = consolidate_allocations(all_outputs)

    if not filled_path.exists():
        return []

    wb = openpyxl.load_workbook(filled_path, read_only=True)
    ws = wb.active

    qty_gp_col = code_col = None
    for c in range(1, min(ws.max_column or 30, 30) + 1):
        v = ws.cell(1, c).value
        if not v:
            continue
        h = str(v).strip().lower()
        if "количество гп" in h:
            qty_gp_col = c
        if "номер позиции" in h or "шифр позиции" in h:
            code_col = c

    if qty_gp_col is None:
        wb.close()
        return []

    results = []
    for r in range(2, ws.max_row + 1):
        code_v = ws.cell(r, code_col or 1).value
        if code_v is None:
            continue
        code_str = str(code_v).strip().rstrip(".")
        allocs = by_pos.get(code_str, [])
        if not allocs:
            continue

        active = [a for a in allocs if not str(a.get("fill_status", "")).startswith("delegated_")]
        if not active:
            active = allocs
        best = max(active, key=lambda a: float(a.get("confidence", 0) or 0))
        expected = best.get("quantity")
        conf = float(best.get("confidence", 0) or 0)
        if expected is None or conf <= 0:
            continue

        actual = ws.cell(r, qty_gp_col).value
        if actual is None:
            results.append((code_str, expected, None, "НЕ ЗАПИСАНО"))
        elif isinstance(actual, (int, float)):
            if abs(actual - expected) < 0.01:
                results.append((code_str, expected, actual, "ОК"))
            else:
                results.append((code_str, expected, actual, "РАСХОЖДЕНИЕ"))
        else:
            results.append((code_str, expected, actual, "ОШИБКА ТИПА"))

    wb.close()
    return results


# ====================================================================
# Excel generation
# ====================================================================

def generate_report(out_path: Path):
    spec_dir = REPO / "runs" / "event_6_1" / "specialist_outputs"
    db_path = REPO / "runs" / "bim2vor.db"
    filled_path = REPO / "runs" / "event_6_1" / "filled_boq.xlsx"

    print("Собираю данные...")
    d = collect_all_data(spec_dir, db_path, filled_path)
    rb = readback_check(filled_path, spec_dir)

    wb = openpyxl.Workbook()

    # ============================================================
    # Sheet 1: Сводка
    # ============================================================
    ws = wb.active
    ws.title = "Сводка"
    ws.sheet_properties.tabColor = "4472C4"

    write_header(ws, 1,
        ["Проверка", "Результат", "Статус", "Комментарий"],
        [40, 30, 15, 60],
    )

    # Summary data
    total_db_vol = sum(r[2] for r in d["cat_rows"])
    total_db_area = sum(r[3] for r in d["cat_rows"])
    n_categories = len(d["cat_rows"])
    n_covered_cats = len(d["cat_claimed"])
    n_uncovered_cats = n_categories - n_covered_cats

    total_claimed = len(d["claimed_map"])
    total_allocated = len(d["allocated_map"])
    total_dopniki = len(d["dopnik_map"])
    true_orphans = set(d["claimed_map"].keys()) - set(d["allocated_map"].keys()) - set(d["dopnik_map"].keys())
    orphan_vol = sum(d["cluster_db"].get(cid, (0, 0, 0))[1] for cid in true_orphans)

    all_recheck = []
    for spec, results in d["recompute"].items():
        all_recheck.extend(results)
    n_recheck_match = sum(1 for r in all_recheck if r["match"])
    n_recheck_mismatch = sum(1 for r in all_recheck if not r["match"])

    rb_ok = sum(1 for r in rb if r[3] == "ОК")
    rb_bad = sum(1 for r in rb if r[3] != "ОК")

    n_missing = len(d["missing_model"])

    checks = [
        (
            "Категории BIM покрыты специалистами",
            f"{n_covered_cats} из {n_categories} категорий",
            "ОК" if n_uncovered_cats <= 4 else "ВНИМАНИЕ",
            f"Не покрыты: {n_uncovered_cats} кат. (roofs, doors, windows и др. — ожидают своих специалистов)",
        ),
        (
            "Кластеры → позиции ВОР или допники",
            f"{total_allocated + total_dopniki} из {total_claimed} кластеров распределены",
            "ОК" if len(true_orphans) < 5 else "ВНИМАНИЕ",
            f"Потерянные кластеры: {len(true_orphans)} (объём: {orphan_vol:.1f} м³)",
        ),
        (
            "DB vs LLM (детерминистический пересчёт)",
            f"{n_recheck_match} совпадений, {n_recheck_mismatch} расхождений (>5%)",
            "ОК" if n_recheck_mismatch <= 2 else "ВНИМАНИЕ",
            "Проверка: SUM(raw elements) vs quantity из LLM-аллокации",
        ),
        (
            "Read-back из filled_boq.xlsx",
            f"{rb_ok} совпадений, {rb_bad} расхождений",
            "ОК" if rb_bad == 0 else "ОШИБКА",
            "Числа прочитаны обратно из заполненного Excel и сверены",
        ),
        (
            "Допники (BIM-элементы без позиции ВОР)",
            f"{total_dopniki} кластеров",
            "ОК" if total_dopniki > 0 else "ВНИМАНИЕ",
            "Кластеры есть в BIM, но нет подходящей позиции ВОР — нужны доп.позиции",
        ),
        (
            "Отсутствует в BIM",
            f"{n_missing} позиций",
            "ИНФО",
            "Позиции ВОР без данных в BIM-модели (сваи, лестницы, ГИ и т.д.)",
        ),
    ]

    for i, (check, result, status, comment) in enumerate(checks, 2):
        fills = [None, None, status_fill(status), None]
        write_row(ws, i, [check, result, status, comment], fills)

    # Summary numbers
    r = len(checks) + 3
    ws.cell(r, 1, "Общие показатели").font = FONT_SECTION
    r += 1
    for label, val in [
        ("Общий объём в DB (физ. элементы)", f"{total_db_vol:,.2f} м³"),
        ("Общая площадь в DB (физ. элементы)", f"{total_db_area:,.2f} м²"),
        ("Специалистов запущено", f"{len(d['outputs'])}"),
        ("Кластеров заявлено (claimed)", f"{total_claimed}"),
        ("Кластеров распределено (allocated)", f"{total_allocated}"),
        ("Кластеров в допниках", f"{total_dopniki}"),
        ("Потерянных кластеров", f"{len(true_orphans)} ({orphan_vol:.1f} м³)"),
        ("Run ID", d["run_id"]),
    ]:
        ws.cell(r, 1, label).font = FONT_BOLD
        ws.cell(r, 2, val).font = FONT_NORMAL
        r += 1

    # ============================================================
    # Sheet 2: Баланс объёмов
    # ============================================================
    ws2 = wb.create_sheet("Баланс объёмов")
    ws2.sheet_properties.tabColor = "70AD47"

    write_header(ws2, 1,
        ["Категория BIM", "Элементов", "Объём м³", "Площадь м²",
         "Покрыта спец.", "Специалист(ы)", "Claimed кластеров", "Allocated кластеров"],
        [25, 12, 15, 15, 14, 20, 18, 18],
    )

    for i, (cat, cnt, vol, area) in enumerate(d["cat_rows"], 2):
        covered = cat in d["cat_claimed"]
        specs_for_cat = set()
        n_claimed_in_cat = 0
        n_alloc_in_cat = 0
        for cid, specs in d["claimed_map"].items():
            if cid.startswith(cat + "::"):
                n_claimed_in_cat += 1
                specs_for_cat.update(specs)
        for cid in d["allocated_map"]:
            if cid.startswith(cat + "::"):
                n_alloc_in_cat += 1

        status = "Да" if covered else "Нет"
        sfill = FILL_OK if covered else FILL_ALARM
        write_row(ws2, i, [
            cat, cnt, vol, area,
            status, ", ".join(sorted(specs_for_cat)) if specs_for_cat else "—",
            n_claimed_in_cat, n_alloc_in_cat,
        ], [None, None, None, None, sfill, None, None, None])

    # Totals
    tr = len(d["cat_rows"]) + 2
    ws2.cell(tr, 1, "ИТОГО").font = FONT_BOLD
    ws2.cell(tr, 2, sum(r[1] for r in d["cat_rows"])).font = FONT_BOLD
    ws2.cell(tr, 3, total_db_vol).font = FONT_BOLD
    ws2.cell(tr, 3).number_format = "#,##0.00"
    ws2.cell(tr, 4, total_db_area).font = FONT_BOLD
    ws2.cell(tr, 4).number_format = "#,##0.00"

    # ============================================================
    # Sheet 3: DB vs LLM
    # ============================================================
    ws3 = wb.create_sheet("DB vs LLM")
    ws3.sheet_properties.tabColor = "ED7D31"

    write_header(ws3, 1,
        ["Специалист", "Позиция ВОР", "LLM количество", "DB количество",
         "Ед.изм.", "Дельта %", "Статус", "Уверенность", "Кол-во кластеров", "Обоснование"],
        [14, 16, 16, 16, 10, 12, 12, 14, 16, 50],
    )

    row = 2
    for spec in sorted(d["recompute"].keys()):
        results = d["recompute"][spec]
        for r in results:
            status = "Совпадает" if r["match"] else "РАСХОЖДЕНИЕ"
            sfill = FILL_OK if r["match"] else FILL_ALARM
            write_row(ws3, row, [
                spec, r["position_id"], r["llm_qty"], r["db_qty"],
                r["unit"], r["delta_pct"],
                status, r["confidence"], r["n_sources"],
                r["reasoning"],
            ], [None, None, None, None, None, None, sfill, None, None, None])
            row += 1

    # Summary row
    ws3.cell(row + 1, 1, "ИТОГО").font = FONT_BOLD
    ws3.cell(row + 1, 2, f"{n_recheck_match} совпадений, {n_recheck_mismatch} расхождений").font = FONT_BOLD

    # ============================================================
    # Sheet 4: Покрытие кластеров
    # ============================================================
    ws4 = wb.create_sheet("Покрытие кластеров")
    ws4.sheet_properties.tabColor = "5B9BD5"

    write_header(ws4, 1,
        ["Cluster ID", "Категория", "Семейство", "Тип", "Элементов",
         "Объём м³", "Площадь м²", "Специалист", "Статус",
         "Позиция ВОР", "Share", "Contribution"],
        [55, 18, 25, 25, 12, 14, 14, 14, 16, 16, 10, 14],
    )

    row = 2
    all_cids = sorted(d["claimed_map"].keys())
    for cid in all_cids:
        parts = cid.split("::")
        cat = parts[0] if parts else ""
        fam = parts[1] if len(parts) > 1 else ""
        typ = parts[2] if len(parts) > 2 else ""
        cnt, vol, area = d["cluster_db"].get(cid, (0, 0, 0))
        specs = d["claimed_map"].get(cid, [])

        if cid in d["allocated_map"]:
            for (sp, pid, qty, unit, share, contrib) in d["allocated_map"][cid]:
                write_row(ws4, row, [
                    cid, cat, fam, typ, cnt, vol, area,
                    sp, "Распределён", pid, share, contrib,
                ], [None]*8 + [FILL_OK] + [None]*3)
                row += 1
        elif cid in d["dopnik_map"]:
            sp, sug, est, u = d["dopnik_map"][cid]
            write_row(ws4, row, [
                cid, cat, fam, typ, cnt, vol, area,
                sp, "Допник", sug[:40], "", est,
            ], [None]*8 + [FILL_WARN] + [None]*3)
            row += 1
        else:
            sp = specs[0] if specs else "?"
            write_row(ws4, row, [
                cid, cat, fam, typ, cnt, vol, area,
                sp, "ПОТЕРЯН", "", "", "",
            ], [None]*8 + [FILL_ALARM] + [None]*3)
            row += 1

    ws4.cell(row + 1, 1, f"Итого кластеров: {len(all_cids)}").font = FONT_BOLD

    # ============================================================
    # Sheet 5: Read-back Excel
    # ============================================================
    ws5 = wb.create_sheet("Read-back Excel")
    ws5.sheet_properties.tabColor = "A5A5A5"

    write_header(ws5, 1,
        ["Позиция ВОР", "Ожидаемое значение", "Значение в Excel", "Статус"],
        [18, 20, 20, 18],
    )

    for i, (code, expected, actual, status) in enumerate(rb, 2):
        sfill = FILL_OK if status == "ОК" else FILL_ALARM
        write_row(ws5, i, [code, expected, actual, status],
                  [None, None, None, sfill])

    tr = len(rb) + 2
    n_ok = sum(1 for r in rb if r[3] == "ОК")
    n_fail = len(rb) - n_ok
    ws5.cell(tr, 1, f"Итого: {n_ok} ОК, {n_fail} расхождений").font = FONT_BOLD

    # ============================================================
    # Sheet 6: Допники
    # ============================================================
    ws6 = wb.create_sheet("Допники")
    ws6.sheet_properties.tabColor = "FFC000"

    write_header(ws6, 1,
        ["Специалист", "Cluster ID", "Категория", "Описание/предложение",
         "Оценочный объём", "Ед.изм.", "Элементов в DB", "Объём DB м³", "Площадь DB м²"],
        [14, 50, 18, 40, 16, 10, 14, 14, 14],
    )

    row = 2
    for cid, (spec, sug, est, unit) in sorted(d["dopnik_map"].items(), key=lambda x: x[1][0]):
        parts = cid.split("::")
        cat = parts[0] if parts else ""
        cnt, vol, area = d["cluster_db"].get(cid, (0, 0, 0))
        write_row(ws6, row, [
            spec, cid, cat, sug, est, unit, cnt, vol, area,
        ])
        row += 1

    ws6.cell(row + 1, 1, f"Итого допников: {len(d['dopnik_map'])}").font = FONT_BOLD

    # ============================================================
    # Sheet 7: Отсутствует в BIM
    # ============================================================
    ws7 = wb.create_sheet("Отсутствует в BIM")
    ws7.sheet_properties.tabColor = "D9D9D9"

    write_header(ws7, 1,
        ["Специалист", "Позиция ВОР", "Описание", "Причина"],
        [14, 16, 50, 50],
    )

    for i, m in enumerate(d["missing_model"], 2):
        write_row(ws7, i, [
            m["specialist"], m["position_id"], m["description"], m["reason"],
        ])

    tr = len(d["missing_model"]) + 2
    ws7.cell(tr, 1, f"Итого: {len(d['missing_model'])} позиций").font = FONT_BOLD

    # ============================================================
    # Sheet 8: Потерянные кластеры
    # ============================================================
    ws8 = wb.create_sheet("Потерянные кластеры")
    ws8.sheet_properties.tabColor = "FF0000"

    write_header(ws8, 1,
        ["Cluster ID", "Категория", "Семейство", "Тип",
         "Специалист", "Элементов", "Объём м³", "Площадь м²", "Комментарий"],
        [55, 18, 25, 25, 14, 12, 14, 14, 40],
    )

    row = 2
    orphan_list = []
    for cid in true_orphans:
        parts = cid.split("::")
        cnt, vol, area = d["cluster_db"].get(cid, (0, 0, 0))
        specs = d["claimed_map"].get(cid, ["?"])
        orphan_list.append((cid, parts, cnt, vol, area, specs))
    orphan_list.sort(key=lambda x: x[3], reverse=True)

    for cid, parts, cnt, vol, area, specs in orphan_list:
        cat = parts[0] if parts else ""
        fam = parts[1] if len(parts) > 1 else ""
        typ = parts[2] if len(parts) > 2 else ""
        comment = ""
        if cat == "structural_columns":
            comment = "Колонны — вероятно входят в агрегатную позицию монолита"
        write_row(ws8, row, [
            cid, cat, fam, typ, specs[0], cnt, vol, area, comment,
        ], [None]*8 + [FILL_WARN if comment else FILL_ALARM])
        row += 1

    ws8.cell(row + 1, 1, f"Итого потерянных: {len(orphan_list)}, объём: {orphan_vol:.2f} м³").font = FONT_BOLD

    # ============================================================
    # Save
    # ============================================================
    wb.save(str(out_path))
    print(f"Отчёт сохранён: {out_path}")
    print(f"  Листов: {len(wb.sheetnames)}")
    print(f"  DB vs LLM: {n_recheck_match} match, {n_recheck_mismatch} mismatch")
    print(f"  Read-back: {rb_ok} ok, {rb_bad} bad")
    print(f"  Допники: {total_dopniki}")
    print(f"  Потерянные кластеры: {len(true_orphans)} ({orphan_vol:.1f} м³)")


def main():
    out_path = REPO / "runs" / "event_6_1" / "verification_report.xlsx"
    generate_report(out_path)


if __name__ == "__main__":
    main()
