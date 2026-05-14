# -*- coding: utf-8 -*-
"""Глубокий анализ ВОР-шаблона с классификацией позиций по типам работ."""
import sys, io, json, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from collections import Counter, defaultdict
from pathlib import Path

FP = Path(r'C:\Users\kuklev.d.s\Downloads\Расчет ПЗ_ВГК№5 (ЖК)_Версия 4.xlsx')
OUT = Path(r'C:\Users\kuklev.d.s\PycharmProjects\bim2vor\data\boq_profile.json')
OUT.parent.mkdir(parents=True, exist_ok=True)

wb = openpyxl.load_workbook(FP, data_only=True)
ws = wb.active
print(f"Лист: {ws.title} ({ws.max_row}×{ws.max_column})")

# Колонки шапки
headers = {}
for c in range(1, ws.max_column + 1):
    v = ws.cell(1, c).value
    if v:
        headers[c] = str(v).strip()
print(f"\nШапка ({len(headers)} колонок):")
for c, h in headers.items():
    print(f"  col {c}: {h}")

# Парсим иерархию по "Шифр позиции" (col 1)
# Примеры: "1.", "1.1.", "1.1.1.", "1.1.2.", "10.2.6.", "194.2"
def parse_code(code: str | None) -> dict | None:
    if not code:
        return None
    s = str(code).strip()
    # Убираем точку в конце
    s_clean = s.rstrip('.')
    parts = s_clean.split('.')
    if not all(p.isdigit() for p in parts):
        # Не стандартный код — возможно текстовый раздел
        return {'raw': s, 'depth': 0, 'parts': []}
    depth = len(parts)
    return {
        'raw': s,
        'parts': [int(p) for p in parts],
        'depth': depth,
        'parent': '.'.join(parts[:-1]) + '.' if depth > 1 else None,
    }


# Обходим все строки
positions = []
for r in range(2, ws.max_row + 1):
    code = ws.cell(r, 1).value  # Шифр
    seq = ws.cell(r, 2).value   # № п/п
    name = ws.cell(r, 6).value  # Наименование
    unit = ws.cell(r, 7).value  # Ед. изм.
    qty = ws.cell(r, 8).value   # Количество подрядное

    if not code and not name:
        continue

    parsed = parse_code(code) if code else None
    positions.append({
        'row': r,
        'code': str(code).strip() if code else None,
        'parsed': parsed,
        'seq': seq,
        'name': str(name).strip() if name else None,
        'unit': str(unit).strip() if unit else None,
        'qty': qty if isinstance(qty, (int, float)) else None,
    })

print(f"\nВсего строк ВОР: {len(positions)}")
# Депт-распределение
depths = Counter(p['parsed']['depth'] if p['parsed'] else 0 for p in positions)
print(f"Глубина иерархии: {dict(depths)}")

# Классификация по ключевым словам
PATTERNS = [
    ('concrete_walls', r'(монтаж|устройство).*бетон.*(стен|колонн)', 'м3'),
    ('concrete_floors', r'(монтаж|устройство).*бетон.*(перекрыт|плит)', 'м3'),
    ('concrete_foundation', r'(монтаж|устройство).*бетон.*(фундамент|плит фундамент)', 'м3'),
    ('formwork_walls', r'опалубк', 'м2'),
    ('rebar', r'(монтаж|устройство).*арматур', 'тн'),
    ('blocks', r'(кладка|устройство).*блок', 'м3'),
    ('brick', r'(кладка|устройство).*кирпич', 'м3'),
    ('insulation', r'(утеплен|изоляц|минерал.*ват)', 'м2'),
    ('plaster', r'штукатур', 'м2'),
    ('screed', r'стяжк', 'м2'),
    ('tile_floor', r'плитк.*(пол|напольн)', 'м2'),
    ('tile_wall', r'плитк.*стен', 'м2'),
    ('paint', r'(окраск|покраск)', 'м2'),
    ('doors', r'(монтаж|установк).*двер', 'шт'),
    ('windows', r'(монтаж|установк).*окн', 'шт'),
    ('elevator', r'лифт', 'шт'),
    ('stairs', r'лестниц', 'м2'),
    ('roof', r'(монтаж|устройство).*кровл', 'м2'),
    ('waterproof', r'(гидроизоляц|водозащ)', 'м2'),
    ('demolition', r'(демонтаж|разборк)', None),
    ('preparation', r'(подготов|устройство).*основан', None),
    ('section_header', r'^(монтажные|общестроительные|внутренн|наружн|инженерн|спецработы|итого)', None),
    ('procurement', r'(приобретен|поставк|комплект)', None),
]

def classify(name: str | None) -> tuple[str, str | None]:
    if not name:
        return 'unknown', None
    nl = name.lower()
    for kind, pat, expected_unit in PATTERNS:
        if re.search(pat, nl):
            return kind, expected_unit
    return 'unknown', None


# Прогоняем
classified = Counter()
unknown_examples = []
for p in positions:
    if p['name']:
        kind, _ = classify(p['name'])
        classified[kind] += 1
        if kind == 'unknown' and p['unit']:
            if len(unknown_examples) < 30:
                unknown_examples.append(p)

print(f"\n=== КЛАССИФИКАЦИЯ ПОЗИЦИЙ ===")
for kind, n in classified.most_common():
    print(f"  {kind:25s} {n:>4}")

print(f"\n=== UNKNOWN с единицей измерения (нужна доразметка) ===")
for p in unknown_examples[:15]:
    print(f"  {p['code']:>10}  {p['unit']:>4}  {p['name'][:80]}")

# Найдём верхнеуровневые разделы (depth=1)
print(f"\n=== ВЕРХНЕУРОВНЕВЫЕ РАЗДЕЛЫ (depth=1) ===")
for p in positions:
    if p['parsed'] and p['parsed']['depth'] == 1:
        print(f"  {p['code']:>6}  {p['name']}")

# Сохраняем профиль
output = {
    'source_file': str(FP),
    'rows_total': len(positions),
    'depth_distribution': dict(depths),
    'classification_summary': dict(classified),
    'positions': positions,
}
OUT.write_text(json.dumps(output, ensure_ascii=False, indent=2, default=str), encoding='utf-8')
print(f"\nПрофиль ВОР: {OUT}")
