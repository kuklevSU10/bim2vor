# -*- coding: utf-8 -*-
"""
Финальное заполнение ВОР: вписывает BIM-количества в колонку 12 (Количество ГП).

Логика:
1. Берём аллокации из specialist_outputs
2. Для каждой аллокации проверяем DB пересчёт
3. Если LLM скопировал заказчика (BIM = Customer) и DB disagrees > 5% → используем DB
4. Если LLM скопировал заказчика и нет source_clusters → не заполняем
5. qty=0 или conf=0 → не заполняем
6. Фасады 06.02.01.08: LLM выдал full cluster — подменяем share-корректным значением
"""
from __future__ import annotations

import json
import sqlite3
import sys
import io
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

REPO = Path(__file__).resolve().parents[1]
FILL_OK = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_WARN = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_ALARM = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def load_customer_quantities(tmpl_path):
    wb = openpyxl.load_workbook(tmpl_path, read_only=True)
    ws = wb.active
    cust = {}
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if code:
            v = ws.cell(r, 9).value
            if isinstance(v, (int, float)):
                cust[str(code).strip().rstrip(".")] = v
    wb.close()
    return cust


def db_recompute(conn, run_id, alloc):
    unit = str(alloc.get("unit", "")).lower().strip()
    sources = alloc.get("source_clusters", [])
    if not sources:
        return None

    total = 0.0
    for sc in sources:
        cid = sc.get("cluster_id", "")
        share = float(sc.get("share", 1.0))
        parts = cid.split("::")
        if len(parts) < 3:
            continue
        row = conn.execute("""
            SELECT COUNT(*), SUM(COALESCE(volume_m3,0)),
                   SUM(COALESCE(area_m2,0)), SUM(COALESCE(length_m,0))
            FROM elements WHERE run_id=? AND category=?
            AND COALESCE(family,'')=? AND COALESCE(type_name,'')=?
            AND is_excluded=0
            AND (is_physical=1 OR category IN ('rooms','apartment_type'))
        """, (run_id, parts[0], parts[1], parts[2])).fetchone()
        if not row:
            continue
        cnt, vol, area, length = row
        if "м3" in unit or "м³" in unit:
            total += vol * share
        elif "м2" in unit or "м²" in unit:
            total += area * share
        elif "шт" in unit:
            total += cnt * share
        elif "пог" in unit or "м.п" in unit:
            total += length * share
        elif "тн" in unit or unit == "т":
            total += vol * share * 2.5

    return round(total, 2) if total > 0 else None


def main():
    tmpl = Path(r"C:\Users\kuklev.d.s\Downloads\Расчет ПЗ_Событие 6.1_Версия 2.xlsx")
    spec_dir = REPO / "runs" / "event_6_1" / "specialist_outputs"
    db_path = REPO / "runs" / "bim2vor.db"
    out = REPO / "runs" / "event_6_1" / "vor_comparison.xlsx"

    # 1. Load specialist outputs
    by_pos = {}
    for p in sorted(spec_dir.glob("*.json")):
        data = json.loads(p.read_text(encoding="utf-8"))
        spec = p.stem
        for a in data.get("phase3_allocations", []):
            pid = a.get("position_id", "").strip().rstrip(".")
            if pid:
                a["_specialist"] = spec
                by_pos.setdefault(pid, []).append(a)

    # 2. Customer quantities
    cust = load_customer_quantities(tmpl)

    # 3. DB connection
    conn = sqlite3.connect(str(db_path))
    run_id = conn.execute(
        "SELECT DISTINCT run_id FROM elements ORDER BY run_id DESC LIMIT 1"
    ).fetchone()[0]

    # 4. Open template
    print("Открываю шаблон ВОР...")
    wb = openpyxl.load_workbook(tmpl)
    ws = wb.active

    filled = []
    skipped = []
    replaced_by_db = []
    dropped_no_source = []

    for r in range(2, ws.max_row + 1):
        code_v = ws.cell(r, 1).value
        if code_v is None:
            continue
        code_str = str(code_v).strip().rstrip(".")

        allocs = by_pos.get(code_str, [])
        if not allocs:
            continue

        active = [a for a in allocs if not str(a.get("fill_status", "")).startswith("delegated_")]
        if not active:
            active = allocs
        best = max(active, key=lambda a: float(a.get("confidence", 0) or 0))

        qty = best.get("quantity")
        conf = float(best.get("confidence", 0) or 0)
        unit = best.get("unit", "")
        spec = best.get("_specialist", "?")
        sources = best.get("source_clusters", [])
        has_sources = len(sources) > 0 and any(sc.get("cluster_id") for sc in sources)

        # Skip: no quantity, zero quantity, or zero confidence
        if qty is None or qty == 0 or conf <= 0:
            skipped.append((code_str, "qty=0 or conf=0"))
            continue

        # Sanity check: if BIM qty is >10x customer qty and conf < 0.5 → suspect
        cust_check = cust.get(code_str)
        if cust_check and cust_check > 0 and qty > 0:
            ratio = qty / cust_check
            if ratio > 10 and conf < 0.5:
                skipped.append((code_str, f"sanity: BIM/Cust={ratio:.1f}x, conf={conf}"))
                continue

        # Check if LLM copied customer value
        cust_qty = cust.get(code_str)
        is_copy = False
        if cust_qty is not None and cust_qty > 0:
            if abs(qty - cust_qty) / cust_qty < 0.001:
                is_copy = True

        # DB verification
        db_qty = db_recompute(conn, run_id, best)
        final_qty = qty
        source_note = ""

        if is_copy:
            if not has_sources:
                dropped_no_source.append((code_str, qty, spec))
                skipped.append((code_str, "copied_no_source"))
                continue

            if db_qty is not None:
                delta_pct = abs(qty - db_qty) / db_qty * 100 if db_qty > 0 else 999
                if delta_pct > 5:
                    final_qty = db_qty
                    source_note = f"DB replace (LLM={qty:.2f}, DB={db_qty:.2f})"
                    replaced_by_db.append((code_str, qty, db_qty, delta_pct, spec))
                    conf = max(0.3, conf - 0.1)
                else:
                    source_note = "DB confirms"

        # Write to col 12
        ws.cell(r, 12).value = round(final_qty, 2)

        if conf >= 0.65:
            ws.cell(r, 12).fill = FILL_OK
        elif conf >= 0.4:
            ws.cell(r, 12).fill = FILL_WARN
        else:
            ws.cell(r, 12).fill = FILL_ALARM

        # DB check status
        if db_qty is not None and final_qty > 0:
            db_delta = abs(final_qty - db_qty) / db_qty * 100 if db_qty > 0 else 0
            db_ok = db_delta < 5
        else:
            db_ok = None
            db_delta = None

        filled.append({
            "row": r,
            "pos": code_str,
            "qty": round(final_qty, 2),
            "db_qty": db_qty,
            "cust": cust_qty,
            "unit": unit,
            "conf": conf,
            "spec": spec,
            "db_ok": db_ok,
            "db_delta": db_delta,
            "note": source_note,
        })

    wb.save(str(out))
    conn.close()

    # Report
    print(f"\nСохранён: {out}")
    print(f"\n{'═' * 80}")
    print(f"РЕЗУЛЬТАТ ЗАПОЛНЕНИЯ")
    print(f"{'═' * 80}")
    print(f"  Заполнено позиций: {len(filled)}")
    print(f"  Пропущено (qty=0/conf=0): {sum(1 for s in skipped if s[1] == 'qty=0 or conf=0')}")
    print(f"  Выброшено (копия заказчика без source): {len(dropped_no_source)}")
    print(f"  Заменено на DB (LLM скопировал заказчика): {len(replaced_by_db)}")
    print()

    if replaced_by_db:
        print("ЗАМЕНЕНЫ НА DB-ЗНАЧЕНИЕ (LLM скопировал заказчика):")
        for pos, llm, db, delta, spec in replaced_by_db:
            print(f"  {pos:18s} LLM={llm:10.2f} → DB={db:10.2f}  (delta={delta:.1f}%)")
        print()

    if dropped_no_source:
        print("НЕ ЗАПОЛНЕНЫ (копия заказчика, нет source_clusters):")
        for pos, qty, spec in dropped_no_source:
            print(f"  {pos:18s} qty={qty:10.2f} spec={spec}")
        print()

    # Full table
    n_db_ok = sum(1 for f in filled if f["db_ok"] is True)
    n_db_bad = sum(1 for f in filled if f["db_ok"] is False)
    n_db_na = sum(1 for f in filled if f["db_ok"] is None)

    print(f"DB ВЕРИФИКАЦИЯ: {n_db_ok} OK, {n_db_bad} расхождений, {n_db_na} без source")
    print()

    print(f"{'Позиция':18s} {'Наше кол-во':>12s} {'DB':>12s} {'Заказчик':>12s} {'Ед':>6s} {'Conf':>6s} {'DB?':>6s} {'Спец':<10s} {'Примечание':<30s}")
    print("-" * 120)
    for f in filled:
        db_str = f"{f['db_qty']:.2f}" if f["db_qty"] else "—"
        cust_str = f"{f['cust']:.2f}" if f["cust"] else "—"
        db_status = "OK" if f["db_ok"] is True else ("FAIL" if f["db_ok"] is False else "—")
        print(f"{f['pos']:18s} {f['qty']:12.2f} {db_str:>12s} {cust_str:>12s} {f['unit'] or '':>6s} {f['conf']:6.2f} {db_status:>6s} {f['spec']:<10s} {f['note']:<30s}")

    print(f"\n{'═' * 80}")
    print(f"ПО СПЕЦИАЛИСТАМ:")
    by_spec = {}
    for f in filled:
        by_spec.setdefault(f["spec"], []).append(f)
    for s, items in sorted(by_spec.items()):
        h = sum(1 for f in items if f["conf"] >= 0.65)
        m = sum(1 for f in items if 0.4 <= f["conf"] < 0.65)
        l = sum(1 for f in items if f["conf"] < 0.4)
        print(f"  {s:15s}: {len(items):3d} позиций (высокая={h}, средняя={m}, низкая={l})")


if __name__ == "__main__":
    main()
